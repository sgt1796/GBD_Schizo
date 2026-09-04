from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from datetime import date
from itertools import combinations, permutations
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

import apc_analysis as apc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BURDEN = ROOT / "prepared_inputs" / "cause_all.csv"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "results"

LOCATIONS = ("China", "United States of America")
SEXES = ("Female", "Male")
OUTCOMES = ("Incidence", "Prevalence", "DALYs")
PROVISIONAL_DECOMPOSITION_AGES = (
    "0-14 years", "15-19 years", "20-24 years", "25-29 years", "30-34 years",
    "35-39 years", "40-44 years", "45-49 years", "50-54 years",
    "55-59 years", "60-64 years", "65-69 years", "70+ years",
)
FINE_DECOMPOSITION_AGES = (
    "0-4 years", "5-9 years", "10-14 years", "15-19 years", "20-24 years",
    "25-29 years", "30-34 years", "35-39 years", "40-44 years",
    "45-49 years", "50-54 years", "55-59 years", "60-64 years",
    "65-69 years", "70-74 years", "75-79 years", "80-84 years",
    "85-89 years", "90-94 years", "95+ years",
)
# Backward-compatible name for code/tests that inspect the current provisional input.
DECOMPOSITION_AGES = PROVISIONAL_DECOMPOSITION_AGES
APC_AGES = apc.BASE_APC_AGES
ALL_AGES = "All ages"
ASR = "Age-standardized"
YEARS = tuple(range(1990, 2024))
ENDPOINTS = (1990, 2023)
ALL_AGE_RECONSTRUCTION_TOLERANCE_PCT = 1e-4
YLD_DALY_IDENTITY_TOLERANCE = 1e-7
PRACTICAL_STABILITY_THRESHOLD_PCT_PER_YEAR = 0.05
PRACTICAL_STABILITY_SENSITIVITY_THRESHOLDS = (0.02, 0.05, 0.10)
COLORS = {"China": "#0072B2", "United States of America": "#D55E00"}
SEX_LINE = {"Female": "-", "Male": "--"}
COMPONENT_COLORS = {
    "population_size_change": "#009E73",
    "age_structure_change": "#E69F00",
    "age_specific_rate_change": "#CC79A7",
}

DESCRIPTIVE_INFERENCE_NOTE = (
    "Descriptive analysis of annual GBD posterior means. Primary analyses report no "
    "hypothesis tests or confidence intervals because posterior draws and cross-year/"
    "cross-stratum correlations are unavailable."
)

OBSOLETE_TABLE_CSVS = frozenset({
    "apc_excluding_2020_2023.csv",
    "apc_excluding_2019_2023.csv",
    "apc_period_curvature.csv",
    "apc_cohort_curvature.csv",
    "apc_summary.csv",
    "apc_local_drift.csv",
    "apc_age_curve.csv",
    "apc_period_rr.csv",
    "apc_cohort_rr.csv",
    "apc_cells.csv",
    "apc_sensitivity_summary_1990_2019.csv",
    "apc_sensitivity_local_drift_1990_2019.csv",
    "apc_sensitivity_period_rr_1990_2019.csv",
    "apc_sensitivity_cohort_rr_1990_2019.csv",
    "pairwise_parallelism.csv",
    "prepandemic_trend_sensitivity.csv",
})
OPTIONAL_NCI_TABLES = frozenset({
    "nci_validation_summary",
    "nci_validation_segments",
    "nci_validation_comparisons",
    "nci_validation_fitted",
})


def clean_measure(value: str) -> str:
    return {
        "DALYs (Disability-Adjusted Life Years)": "DALYs",
        "YLDs (Years Lived with Disability)": "YLDs",
    }.get(str(value), str(value))


def load_burden(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, low_memory=False)
    required = {
        "location_name", "sex_name", "age_name", "measure_name", "metric_name",
        "cause_name", "year", "val", "lower", "upper",
    }
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Burden file is missing columns: {sorted(missing)}")
    df = df.copy()
    causes = set(df["cause_name"].dropna().astype(str).str.strip())
    if causes != {"Schizophrenia"}:
        raise ValueError(f"Burden file must contain only cause_name='Schizophrenia'; found {sorted(causes)}")
    df["measure_name"] = df["measure_name"].map(clean_measure)
    for col in ("year", "val", "lower", "upper"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df[
        df["location_name"].isin(LOCATIONS)
        & df["sex_name"].isin(SEXES)
        & df["year"].isin(YEARS)
        & df["measure_name"].isin((*OUTCOMES, "YLDs"))
        & df["metric_name"].isin(("Number", "Rate"))
    ].copy()
    return df


def _normalise_population_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "location": "location_name", "sex": "sex_name", "age": "age_name",
        "population": "population", "pop": "population", "value": "population",
        "val": "population",
    }
    rename = {}
    for col in df.columns:
        key = col.strip().lower()
        if key in aliases and aliases[key] not in df.columns:
            rename[col] = aliases[key]
    return df.rename(columns=rename)


def select_decomposition_ages(available_ages: set[str]) -> tuple[str, ...]:
    """Select the complete finest supported age partition."""
    if set(FINE_DECOMPOSITION_AGES) <= available_ages:
        return FINE_DECOMPOSITION_AGES
    if set(PROVISIONAL_DECOMPOSITION_AGES) <= available_ages:
        return PROVISIONAL_DECOMPOSITION_AGES
    fine_missing = sorted(set(FINE_DECOMPOSITION_AGES) - available_ages)
    provisional_missing = sorted(set(PROVISIONAL_DECOMPOSITION_AGES) - available_ages)
    raise ValueError(
        "No complete supported decomposition age partition was found. "
        f"Missing from fine partition: {fine_missing}; missing from provisional "
        f"partition: {provisional_missing}."
    )


def burden_age_granularity_audit(df: pd.DataFrame) -> pd.DataFrame:
    """Record whether production-grade fine-age burden panels are complete."""
    rows = []
    keys = ["location_name", "sex_name", "age_name", "year"]
    for outcome in OUTCOMES:
        panel = df[
            df.measure_name.eq(outcome)
            & df.metric_name.isin(("Number", "Rate"))
            & df.age_name.isin(FINE_DECOMPOSITION_AGES)
        ]
        expected_per_metric = (
            len(LOCATIONS) * len(SEXES) * len(FINE_DECOMPOSITION_AGES) * len(YEARS)
        )
        number = panel[panel.metric_name.eq("Number")]
        rate = panel[panel.metric_name.eq("Rate")]
        rows.append(
            {
                "measure_name": outcome,
                "required_age_groups": len(FINE_DECOMPOSITION_AGES),
                "available_required_age_groups": int(
                    panel.age_name.nunique()
                ),
                "expected_cells_per_metric": expected_per_metric,
                "number_cells": int(len(number)),
                "rate_cells": int(len(rate)),
                "duplicate_number_keys": int(number.duplicated(keys).sum()),
                "duplicate_rate_keys": int(rate.duplicated(keys).sum()),
                "fine_age_panel_complete": bool(
                    len(number) == expected_per_metric
                    and len(rate) == expected_per_metric
                    and not number.duplicated(keys).any()
                    and not rate.duplicated(keys).any()
                ),
            }
        )
    return pd.DataFrame(rows)


def validate_required_burden_summaries(df: pd.DataFrame) -> None:
    """Require the all-age counts and ASRs used by the frozen primary analysis."""
    keys = ["location_name", "sex_name", "measure_name", "age_name", "metric_name", "year"]
    required = (
        (ALL_AGES, "Number", "all-age Number"),
        (ASR, "Rate", "age-standardized Rate"),
    )
    expected = len(LOCATIONS) * len(SEXES) * len(OUTCOMES) * len(YEARS)
    problems = []
    for age_name, metric_name, label in required:
        panel = df[
            df.measure_name.isin(OUTCOMES)
            & df.age_name.eq(age_name)
            & df.metric_name.eq(metric_name)
        ]
        duplicate_count = int(panel.duplicated(keys).sum())
        if len(panel) != expected or duplicate_count:
            problems.append(
                f"{label}: found {len(panel)} rows and {duplicate_count} duplicates; "
                f"expected {expected} unique rows"
            )
    if problems:
        raise ValueError(
            "Burden input is incomplete for the primary descriptive analysis: "
            + "; ".join(problems)
        )


def load_official_population(path: Path, release: str) -> pd.DataFrame:
    df = _normalise_population_columns(pd.read_csv(path, low_memory=False))
    required = {"location_name", "sex_name", "age_name", "year", "population", "gbd_release"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Population file is missing columns: {sorted(missing)}")
    if "2023" not in str(release).lower():
        raise ValueError("--population-release must explicitly identify the GBD 2023 release.")
    file_releases = df["gbd_release"].dropna().astype(str).str.strip()
    file_releases = file_releases[file_releases.ne("")]
    if file_releases.empty or not file_releases.str.contains("2023", regex=False).all():
        found = sorted(file_releases.unique().tolist())
        raise ValueError(f"Population file gbd_release values must consistently identify GBD 2023; found {found}")
    out = df[
        df["location_name"].isin(LOCATIONS)
        & df["sex_name"].isin(SEXES)
        & df["age_name"].isin(FINE_DECOMPOSITION_AGES)
        & pd.to_numeric(df["year"], errors="coerce").isin(YEARS)
    ].copy()
    out["year"] = pd.to_numeric(out["year"], errors="raise").astype(int)
    out["population"] = pd.to_numeric(out["population"], errors="raise")
    out["population_source"] = "official_GBD_2023"
    out = out[["location_name", "sex_name", "age_name", "year", "population", "population_source"]]
    validate_population(out, required_ages=FINE_DECOMPOSITION_AGES)
    return out


def infer_proxy_population(
    df: pd.DataFrame, allow_undefined: bool = False
) -> pd.DataFrame:
    ages = select_decomposition_ages(set(df.age_name.dropna().astype(str)))
    keys = ["location_name", "sex_name", "age_name", "year", "measure_name"]
    num = df[(df.metric_name == "Number") & df.age_name.isin(ages)][keys + ["val"]].rename(columns={"val": "number"})
    rate = df[(df.metric_name == "Rate") & df.age_name.isin(ages)][keys + ["val"]].rename(columns={"val": "rate"})
    merged = num.merge(rate, on=keys, validate="one_to_one")
    invalid = merged.rate.eq(0) & merged.number.ne(0)
    if invalid.any():
        raise ValueError("Cannot reconstruct population where Rate is zero but Number is nonzero.")
    merged["population"] = np.where(
        merged.rate.gt(0), merged.number / merged.rate * 100000.0, np.nan
    )
    out = merged.groupby(keys[:-1], as_index=False, dropna=False)["population"].median()

    expected_index = pd.MultiIndex.from_product(
        [LOCATIONS, SEXES, ages, YEARS], names=keys[:-1]
    )
    out = (
        out.set_index(keys[:-1])
        .reindex(expected_index)
        .reset_index()
    )
    missing = out.population.isna()
    if missing.any():
        # Some disorders legitimately have Number=Rate=0 in a fine age group,
        # making the direct ratio undefined. If exactly one age is undefined,
        # reconstruct total population from exported All ages Number/Rate and
        # fill the residual. This remains an explicitly nonofficial proxy.
        total_keys = ["location_name", "sex_name", "year", "measure_name"]
        all_number = df[
            df.age_name.eq(ALL_AGES) & df.metric_name.eq("Number")
        ][total_keys + ["val"]].rename(columns={"val": "number"})
        all_rate = df[
            df.age_name.eq(ALL_AGES) & df.metric_name.eq("Rate")
        ][total_keys + ["val"]].rename(columns={"val": "rate"})
        totals = all_number.merge(all_rate, on=total_keys, validate="one_to_one")
        totals = totals[totals.rate.gt(0)].copy()
        totals["total_population"] = totals.number / totals.rate * 100000.0
        totals = totals.groupby(total_keys[:-1], as_index=False).total_population.median()
        totals = totals.set_index(total_keys[:-1]).total_population

        group_keys = ["location_name", "sex_name", "year"]
        for group, indices in out.groupby(group_keys, sort=False).groups.items():
            indices = list(indices)
            missing_indices = [index for index in indices if pd.isna(out.at[index, "population"])]
            if not missing_indices:
                continue
            if len(missing_indices) != 1 or group not in totals.index:
                if allow_undefined:
                    continue
                missing_ages = out.loc[missing_indices, "age_name"].tolist()
                raise ValueError(
                    "Cannot reconstruct proxy population for undefined Number/Rate cells: "
                    f"group={group}, missing_ages={missing_ages}."
                )
            residual = float(totals.loc[group]) - float(
                out.loc[indices, "population"].sum(skipna=True)
            )
            if not np.isfinite(residual) or residual <= 0:
                raise ValueError(
                    "All-age residual population is not finite and positive for "
                    f"group={group}: {residual}."
                )
            out.at[missing_indices[0], "population"] = residual

    out["population_source"] = "derived_proxy_NOT_OFFICIAL"
    return out


def validate_population(
    pop: pd.DataFrame, required_ages: tuple[str, ...] | None = None
) -> None:
    keys = ["location_name", "sex_name", "age_name", "year"]
    if pop.duplicated(keys).any():
        raise ValueError("Population file contains duplicate dimensional keys.")
    ages = required_ages or select_decomposition_ages(
        set(pop.age_name.dropna().astype(str))
    )
    if set(pop.age_name) != set(ages):
        unexpected = sorted(set(pop.age_name) - set(ages))
        missing = sorted(set(ages) - set(pop.age_name))
        raise ValueError(
            f"Population ages do not match the required partition; missing={missing}, "
            f"unexpected={unexpected}."
        )
    expected = len(LOCATIONS) * len(SEXES) * len(ages) * len(YEARS)
    if len(pop) != expected:
        raise ValueError(f"Population file has {len(pop)} rows; expected {expected} complete rows.")
    if not np.isfinite(pop["population"]).all() or (pop["population"] <= 0).any():
        raise ValueError("Population values must be finite and positive.")


def compare_population_sources(
    official: pd.DataFrame, reconstructed: pd.DataFrame
) -> pd.DataFrame:
    """Record official-versus-count/rate population discrepancies cell by cell."""
    keys = ["location_name", "sex_name", "age_name", "year"]
    left = official[keys + ["population"]].rename(
        columns={"population": "official_population"}
    )
    right = reconstructed[keys + ["population"]].rename(
        columns={"population": "reconstructed_population"}
    )
    out = left.merge(right, on=keys, how="outer", validate="one_to_one", indicator=True)
    out["absolute_difference"] = (
        out.official_population - out.reconstructed_population
    )
    out["relative_difference_pct"] = (
        100.0 * out.absolute_difference / out.official_population
    )
    out["key_match_status"] = out.pop("_merge").astype(str)
    out["reconstruction_available"] = np.isfinite(out.reconstructed_population)
    out["comparison_status"] = np.where(
        out.reconstruction_available,
        "compared",
        "unavailable_zero_burden_number_and_rate",
    )
    return out


def audit_burden(df: pd.DataFrame, pop: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    ages = select_decomposition_ages(set(pop.age_name.dropna().astype(str)))
    key_cols = ["location_name", "sex_name", "age_name", "measure_name", "metric_name", "year"]
    rows = []
    for outcome in OUTCOMES:
        for loc in LOCATIONS:
            for sex in SEXES:
                panel = df[(df.measure_name == outcome) & (df.location_name == loc) & (df.sex_name == sex)]
                counts = panel[(panel.metric_name == "Number") & (panel.age_name == ALL_AGES)]
                rates = panel[(panel.metric_name == "Rate") & (panel.age_name == ASR)]
                age_groups = panel[panel.age_name.isin(ages)]
                rows.append({
                    "location_name": loc, "sex_name": sex, "measure_name": outcome,
                    "all_age_count_years": counts.year.nunique(), "asr_years": rates.year.nunique(),
                    "all_age_groups": age_groups.age_name.nunique(),
                    "all_age_year_cells": age_groups[["age_name", "year"]].drop_duplicates().shape[0],
                    "missing_values": int(panel[["val", "lower", "upper"]].isna().sum().sum()),
                    "invalid_ui_rows": int(((panel.lower > panel.val) | (panel.val > panel.upper)).sum()),
                    "negative_rows": int((panel.val < 0).sum()),
                    "zero_rows": int((panel.val == 0).sum()),
                    "nonpositive_rows": int((panel.val <= 0).sum()),
                })
    audit = pd.DataFrame(rows)
    duplicate_audit = pd.DataFrame([{
        "rows": len(df), "duplicate_dimensional_keys": int(df.duplicated(key_cols).sum()),
        "population_rows": len(pop),
        "population_source": pop.population_source.iloc[0],
    }])

    # Reconstruction check uses official population when available; otherwise it is explicitly provisional.
    age_rates = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & df.age_name.isin(ages)][
        ["location_name", "sex_name", "age_name", "measure_name", "year", "val"]
    ].rename(columns={"val": "rate"})
    age_counts = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Number") & df.age_name.isin(ages)][
        ["location_name", "sex_name", "age_name", "measure_name", "year", "val"]
    ].rename(columns={"val": "reported_count"})
    recon = age_rates.merge(pop, on=["location_name", "sex_name", "age_name", "year"], validate="many_to_one")
    recon = recon.merge(age_counts, on=["location_name", "sex_name", "age_name", "measure_name", "year"], validate="one_to_one")
    recon["reconstructed_count"] = recon.population * recon.rate / 100000.0
    recon["absolute_error"] = recon.reconstructed_count - recon.reported_count
    recon["relative_error_pct"] = 100.0 * recon.absolute_error / recon.reported_count
    return audit, duplicate_audit, recon


def audit_source_export_zeros(
    df: pd.DataFrame, burden_metadata_path: Path | None
) -> tuple[pd.DataFrame, dict]:
    """Describe exact zero cells and verify their preserved-export provenance."""
    fine = df[
        df.measure_name.isin(OUTCOMES)
        & df.metric_name.isin(("Number", "Rate"))
        & df.age_name.isin(FINE_DECOMPOSITION_AGES)
    ].copy()
    zeros = fine[fine.val.eq(0)].copy()
    summary = (
        zeros.groupby(["measure_name", "age_name", "metric_name"], sort=False)
        .agg(
            zero_cell_count=("val", "size"),
            location_count=("location_name", "nunique"),
            sex_count=("sex_name", "nunique"),
            year_count=("year", "nunique"),
            lower_also_zero=("lower", lambda values: bool(values.eq(0).all())),
            upper_also_zero=("upper", lambda values: bool(values.eq(0).all())),
        )
        .reset_index()
    )
    expected_ages = {
        "Incidence": {
            "0-4 years", "5-9 years", "80-84 years", "85-89 years",
            "90-94 years", "95+ years",
        },
        "Prevalence": {"0-4 years", "5-9 years"},
        "DALYs": {"0-4 years", "5-9 years"},
    }
    observed_ages = {
        outcome: set(zeros.loc[zeros.measure_name.eq(outcome), "age_name"])
        for outcome in OUTCOMES
    }
    expected_pattern = observed_ages == expected_ages
    complete_panel_pattern = bool(
        len(zeros) == 2720
        and summary.zero_cell_count.eq(len(LOCATIONS) * len(SEXES) * len(YEARS)).all()
        and summary.location_count.eq(len(LOCATIONS)).all()
        and summary.sex_count.eq(len(SEXES)).all()
        and summary.year_count.eq(len(YEARS)).all()
    )

    provenance_path = (
        burden_metadata_path.parent / "structural_zero_provenance.json"
        if burden_metadata_path is not None
        else None
    )
    provenance = {}
    if provenance_path is not None and provenance_path.is_file():
        try:
            provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            provenance = {}
    source_verified = bool(
        provenance.get("source_to_canonical_zero_keys_exact_match") is True
        and provenance.get("total_zero_cells") == len(zeros)
    )
    summary["source_export_provenance_verified"] = source_verified
    summary["interpretation"] = (
        "Exact zero is present in the preserved IHME export; the export alone does not "
        "distinguish a biological structural zero from a GBD model-support convention."
    )
    validation = {
        "total_fine_age_zero_cells": int(len(zeros)),
        "expected_age_outcome_pattern": bool(expected_pattern),
        "complete_location_sex_year_metric_pattern": complete_panel_pattern,
        "source_export_provenance_verified": source_verified,
        "provenance_file": portable_path(provenance_path) if provenance_path else None,
    }
    return summary, validation


def all_age_count_reconstruction(df: pd.DataFrame, reconstruction: pd.DataFrame) -> pd.DataFrame:
    """Compare summed age-specific reconstructed counts with reported all-age counts."""
    keys = ["location_name", "sex_name", "measure_name", "year"]
    reconstructed = reconstruction.groupby(keys, as_index=False).agg(
        reconstructed_all_age_count=("reconstructed_count", "sum")
    )
    reported = df[
        df.measure_name.isin(OUTCOMES)
        & (df.metric_name == "Number")
        & (df.age_name == ALL_AGES)
    ][keys + ["val", "lower", "upper"]].rename(columns={
        "val": "reported_all_age_count",
        "lower": "reported_all_age_lower",
        "upper": "reported_all_age_upper",
    })
    out = reconstructed.merge(reported, on=keys, validate="one_to_one")
    out["absolute_error"] = out.reconstructed_all_age_count - out.reported_all_age_count
    out["relative_error_pct"] = 100.0 * out.absolute_error / out.reported_all_age_count
    out["within_tolerance"] = out.relative_error_pct.abs() <= ALL_AGE_RECONSTRUCTION_TOLERANCE_PCT
    out["tolerance_pct"] = ALL_AGE_RECONSTRUCTION_TOLERANCE_PCT
    return out


def verify_yld_daly_identity(df: pd.DataFrame) -> pd.DataFrame:
    """Audit an optional YLD panel without making it a production input.

    Schizophrenia DALYs and YLDs are identical in the provisional source
    exports, so retaining both would duplicate an outcome.  The production
    contract, however, requires incidence, prevalence, and DALYs only.  A
    completely absent YLD panel is therefore a valid ``not_available`` audit
    result, while a partial or non-identical panel fails the audit.
    """
    ages = select_decomposition_ages(set(df.age_name.dropna().astype(str)))
    keys = ["location_name", "sex_name", "age_name", "metric_name", "year"]
    a = df[df.measure_name == "DALYs"][keys + ["val", "lower", "upper"]]
    b = df[df.measure_name == "YLDs"][keys + ["val", "lower", "upper"]]
    expected_age_metrics = {
        *((age, metric) for age in ages for metric in ("Number", "Rate")),
        (ALL_AGES, "Number"), (ALL_AGES, "Rate"), (ASR, "Rate"),
    }
    expected_cells = len(LOCATIONS) * len(SEXES) * len(YEARS) * len(expected_age_metrics)
    duplicate_cells = int(a.duplicated(keys).sum() + b.duplicated(keys).sum())
    if b.empty:
        return pd.DataFrame([{
            "audit_status": "not_available",
            "audit_passed": True,
            "yld_panel_available": False,
            "matched_cells": 0,
            "expected_cells": expected_cells,
            "duplicate_cells": duplicate_cells,
            "complete_expected_panel": False,
            **{
                f"max_{kind}_difference_{field}": np.nan
                for field in ("val", "lower", "upper")
                for kind in ("abs", "relative")
            },
            "identity_tolerance": YLD_DALY_IDENTITY_TOLERANCE,
            "numerically_identical": False,
        }])

    m = a.merge(b, on=keys, suffixes=("_daly", "_yld"), validate="one_to_one")
    observed_age_metrics = set(map(tuple, m[["age_name", "metric_name"]].drop_duplicates().to_numpy()))
    complete = bool(
        len(a) == expected_cells
        and len(b) == expected_cells
        and len(m) == expected_cells
        and duplicate_cells == 0
        and observed_age_metrics == expected_age_metrics
    )
    result = {"matched_cells": len(m), "expected_cells": expected_cells,
              "duplicate_cells": duplicate_cells, "complete_expected_panel": complete}
    for col in ("val", "lower", "upper"):
        delta = np.abs(m[f"{col}_daly"] - m[f"{col}_yld"])
        denom = np.maximum(np.abs(m[f"{col}_daly"]), 1e-12)
        result[f"max_abs_difference_{col}"] = float(delta.max())
        result[f"max_relative_difference_{col}"] = float((delta / denom).max())
    result["identity_tolerance"] = YLD_DALY_IDENTITY_TOLERANCE
    result["numerically_identical"] = bool(
        complete and all(
            result[f"max_relative_difference_{col}"] < YLD_DALY_IDENTITY_TOLERANCE
            for col in ("val", "lower", "upper")
        )
    )
    result["yld_panel_available"] = True
    result["audit_passed"] = result["numerically_identical"]
    result["audit_status"] = (
        "verified_identical"
        if result["numerically_identical"]
        else "incomplete_or_nonidentical"
    )
    return pd.DataFrame([result])


def percent_change(start: float, end: float) -> float:
    return 100.0 * (end / start - 1.0) if start > 0 else np.nan


def endpoint_table(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for loc in LOCATIONS:
        for sex in SEXES:
            for outcome in OUTCOMES:
                p = df[(df.location_name == loc) & (df.sex_name == sex) & (df.measure_name == outcome)]
                for metric, age in (("Number", ALL_AGES), ("Rate", ASR)):
                    s = p[(p.metric_name == metric) & (p.age_name == age)].set_index("year")
                    a, b = s.loc[1990], s.loc[2023]
                    rows.append({
                        "location_name": loc, "sex_name": sex, "measure_name": outcome,
                        "metric_name": "All-age count" if metric == "Number" else "Age-standardized rate per 100,000",
                        "value_1990": a.val, "lower_1990": a.lower, "upper_1990": a.upper,
                        "value_2023": b.val, "lower_2023": b.lower, "upper_2023": b.upper,
                        "absolute_change": b.val - a.val,
                        "percent_change_point_estimate": percent_change(a.val, b.val),
                        "interval_dominance_change": "increase" if b.lower > a.upper else ("decrease" if b.upper < a.lower else "overlap/inconclusive"),
                        "uncertainty_note": "Native endpoint UIs; change is a point estimate, not a GBD draw-derived UI.",
                    })
    return pd.DataFrame(rows)


def contrast_tables(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    asr = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR)]
    country_rows, sex_rows = [], []
    for year in ENDPOINTS:
        for sex in SEXES:
            for outcome in OUTCOMES:
                p = asr[(asr.year == year) & (asr.sex_name == sex) & (asr.measure_name == outcome)].set_index("location_name")
                a, b = p.loc["China"], p.loc["United States of America"]
                country_rows.append({
                    "year": year, "sex_name": sex, "measure_name": outcome,
                    "china_asr": a.val, "us_asr": b.val, "china_us_rate_ratio": a.val / b.val,
                    "absolute_rate_difference": a.val - b.val,
                    "interval_dominance": "China lower" if a.upper < b.lower else ("China higher" if a.lower > b.upper else "overlap/inconclusive"),
                    "derived_uncertainty_note": "Point-estimate ratio/difference; no posterior-draw UI.",
                })
        for loc in LOCATIONS:
            for outcome in OUTCOMES:
                p = asr[(asr.year == year) & (asr.location_name == loc) & (asr.measure_name == outcome)].set_index("sex_name")
                m, f = p.loc["Male"], p.loc["Female"]
                sex_rows.append({
                    "year": year, "location_name": loc, "measure_name": outcome,
                    "male_asr": m.val, "female_asr": f.val, "male_female_rate_ratio": m.val / f.val,
                    "absolute_rate_difference": m.val - f.val,
                    "interval_dominance": "Male higher" if m.lower > f.upper else ("Male lower" if m.upper < f.lower else "overlap/inconclusive"),
                    "derived_uncertainty_note": "Point-estimate ratio/difference; no posterior-draw UI.",
                })
    return pd.DataFrame(country_rows), pd.DataFrame(sex_rows)


def segmented_design(years: np.ndarray, knots: tuple[int, ...]) -> np.ndarray:
    x = years.astype(float) - float(years.min())
    return np.column_stack([np.ones(len(years)), x, *[np.maximum(0.0, years - k) for k in knots]])


def candidate_knots(years: np.ndarray, count: int, min_segment: int = 4) -> list[tuple[int, ...]]:
    if count == 0:
        return [()]
    candidates = [int(x) for x in years if x - years.min() >= min_segment and years.max() - x >= min_segment]
    return [k for k in combinations(candidates, count) if all(b - a >= min_segment for a, b in zip((int(years.min()), *k), (*k, int(years.max()))))]


def fit_linear(y: np.ndarray, X: np.ndarray, weights: np.ndarray | None = None) -> dict:
    if weights is None:
        Xw, yw = X, y
    else:
        root = np.sqrt(np.asarray(weights, float))
        Xw, yw = X * root[:, None], y * root
    beta = np.linalg.pinv(Xw.T @ Xw) @ (Xw.T @ yw)
    residual = y - X @ beta
    rss = float(np.sum(residual**2 if weights is None else weights * residual**2))
    dof = max(len(y) - X.shape[1], 1)
    sigma2 = rss / dof
    cov = sigma2 * np.linalg.pinv(Xw.T @ Xw)
    return {"beta": beta, "residual": residual, "rss": rss, "cov": cov, "dof": dof, "fitted": X @ beta}


def best_segmented_fit(
    years: np.ndarray,
    y: np.ndarray,
    knot_count: int,
    weights: np.ndarray | None = None,
    min_segment: int = 4,
) -> dict:
    best = None
    for knots in candidate_knots(years, knot_count, min_segment=min_segment):
        X = segmented_design(years, knots)
        fit = fit_linear(y, X, weights)
        if best is None or fit["rss"] < best["rss"]:
            best = fit | {"knots": knots, "X": X}
    if best is None:
        raise ValueError("No valid segmented model candidate.")
    return best


def select_segmented_bic(
    years: np.ndarray,
    y: np.ndarray,
    max_joinpoints: int = 2,
    min_segment: int = 4,
) -> dict:
    """Select a descriptive segmented curve using a conservative BIC heuristic.

    A breakpoint contributes both a slope-change coefficient and an estimated
    location to the parameter penalty. BIC is used only to summarize shape; it
    is not treated as formal evidence for a change point.
    """
    candidates = []
    for knot_count in range(max_joinpoints + 1):
        fit = best_segmented_fit(
            years, y, knot_count, min_segment=min_segment
        )
        n = len(y)
        parameter_count = fit["X"].shape[1] + knot_count
        bic = n * math.log(max(fit["rss"] / n, np.finfo(float).tiny)) + parameter_count * math.log(n)
        candidates.append(fit | {"bic": float(bic), "bic_parameter_count": parameter_count})
    selected = min(candidates, key=lambda item: item["bic"])
    selected["candidate_bic"] = {len(item["knots"]): item["bic"] for item in candidates}
    return selected


def fit_ar1_gls(
    y: np.ndarray, X: np.ndarray, max_iterations: int = 100, tolerance: float = 1e-10
) -> dict:
    """Fit a stationary AR(1) feasible-GLS model by Prais-Winsten iteration."""
    y = np.asarray(y, float)
    X = np.asarray(X, float)
    beta = np.linalg.pinv(X) @ y
    rho = 0.0
    for _ in range(max_iterations):
        residual = y - X @ beta
        denominator = float(residual[:-1] @ residual[:-1])
        updated_rho = (
            float(residual[1:] @ residual[:-1] / denominator)
            if denominator > np.finfo(float).tiny
            else 0.0
        )
        updated_rho = float(np.clip(updated_rho, -0.95, 0.95))
        scale0 = math.sqrt(max(1.0 - updated_rho**2, np.finfo(float).eps))
        transformed_y = np.r_[scale0 * y[0], y[1:] - updated_rho * y[:-1]]
        transformed_X = np.vstack(
            [scale0 * X[0], X[1:] - updated_rho * X[:-1]]
        )
        updated_beta = np.linalg.pinv(transformed_X) @ transformed_y
        converged = bool(
            abs(updated_rho - rho) < tolerance
            and np.max(np.abs(updated_beta - beta)) < tolerance
        )
        rho, beta = updated_rho, updated_beta
        if converged:
            break
    residual = y - X @ beta
    innovations = np.r_[
        math.sqrt(max(1.0 - rho**2, np.finfo(float).eps)) * residual[0],
        residual[1:] - rho * residual[:-1],
    ]
    innovation_rss = float(innovations @ innovations)
    return {
        "beta": beta,
        "fitted": X @ beta,
        "residual": residual,
        "rho": rho,
        "innovation_rss": innovation_rss,
        "iterations_converged": converged,
    }


def select_segmented_ar1_bic(
    years: np.ndarray,
    y: np.ndarray,
    max_joinpoints: int = 2,
    min_segment: int = 4,
) -> dict:
    """Select a segmented AR(1) GLS sensitivity using Gaussian likelihood BIC."""
    candidates = []
    n = len(y)
    for knot_count in range(max_joinpoints + 1):
        best = None
        for knots in candidate_knots(years, knot_count, min_segment=min_segment):
            X = segmented_design(years, knots)
            fit = fit_ar1_gls(y, X)
            parameter_count = X.shape[1] + knot_count + 1  # includes AR(1) rho
            bic = (
                n * math.log(max(fit["innovation_rss"] / n, np.finfo(float).tiny))
                - math.log(max(1.0 - fit["rho"] ** 2, np.finfo(float).eps))
                + parameter_count * math.log(n)
            )
            item = fit | {
                "knots": knots,
                "X": X,
                "bic": float(bic),
                "bic_parameter_count": parameter_count,
            }
            if best is None or item["bic"] < best["bic"]:
                best = item
        if best is None:
            raise ValueError("No valid segmented AR(1) model candidate.")
        candidates.append(best)
    selected = min(candidates, key=lambda item: item["bic"])
    selected["candidate_bic"] = {
        len(item["knots"]): item["bic"] for item in candidates
    }
    return selected


def segmented_ar1_sensitivity(
    df: pd.DataFrame, primary_summary: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compare primary segmented curves with AR(1) feasible-GLS sensitivity fits."""
    trend = df[
        df.measure_name.isin(OUTCOMES)
        & df.metric_name.eq("Rate")
        & df.age_name.eq(ASR)
    ]
    primary = primary_summary.set_index(["location_name", "sex_name", "measure_name"])
    summaries = []
    segments = []
    for keys, panel in trend.groupby(
        ["location_name", "sex_name", "measure_name"], sort=True
    ):
        panel = panel.sort_values("year")
        years = panel.year.to_numpy(int)
        selected = select_segmented_ar1_bic(years, np.log(panel.val.to_numpy(float)))
        span = float(years.max() - years.min())
        contrast = np.zeros(len(selected["beta"]))
        contrast[1] = 1.0
        for index, knot in enumerate(selected["knots"]):
            contrast[2 + index] = (years.max() - knot) / span
        aapc = 100.0 * (math.exp(float(contrast @ selected["beta"])) - 1.0)
        primary_row = primary.loc[keys]
        summaries.append(
            {
                "location_name": keys[0],
                "sex_name": keys[1],
                "measure_name": keys[2],
                "primary_joinpoint_count": int(primary_row.joinpoint_count),
                "primary_joinpoint_years": primary_row.joinpoint_years,
                "primary_aapc": float(primary_row.aapc),
                "ar1_joinpoint_count": len(selected["knots"]),
                "ar1_joinpoint_years": ",".join(map(str, selected["knots"])),
                "ar1_aapc": aapc,
                "ar1_minus_primary_aapc": aapc - float(primary_row.aapc),
                "ar1_rho": float(selected["rho"]),
                "ar1_bic": float(selected["bic"]),
                "iterations_converged": bool(selected["iterations_converged"]),
                "primary_practical_label": trend_direction(float(primary_row.aapc)),
                "ar1_practical_label": trend_direction(aapc),
                "practical_label_agreement": trend_direction(float(primary_row.aapc)) == trend_direction(aapc),
                "interpretation": (
                    "AR(1) feasible-GLS robustness analysis of posterior-mean series; "
                    "not GBD posterior uncertainty."
                ),
            }
        )
        breaks = (int(years.min()), *selected["knots"], int(years.max()))
        for segment_index, (start, end) in enumerate(zip(breaks[:-1], breaks[1:]), 1):
            slope_contrast = np.zeros(len(selected["beta"]))
            slope_contrast[1] = 1.0
            for index, knot in enumerate(selected["knots"]):
                slope_contrast[2 + index] = 1.0 if start >= knot else 0.0
            segments.append(
                {
                    "location_name": keys[0],
                    "sex_name": keys[1],
                    "measure_name": keys[2],
                    "segment_index": segment_index,
                    "start_year": start,
                    "end_year": end,
                    "ar1_segment_apc": 100.0 * (
                        math.exp(float(slope_contrast @ selected["beta"])) - 1.0
                    ),
                    "ar1_rho": float(selected["rho"]),
                    "formal_inference_performed": False,
                }
            )
    return pd.DataFrame(summaries), pd.DataFrame(segments)


def residual_diagnostics(residual: np.ndarray) -> dict[str, float | bool]:
    residual = np.asarray(residual, float)
    denominator = float(np.sum(residual**2))
    if len(residual) < 2 or denominator <= np.finfo(float).tiny:
        return {
            "lag1_residual_autocorrelation": np.nan,
            "durbin_watson": np.nan,
            "material_residual_autocorrelation": False,
        }
    left, right = residual[:-1], residual[1:]
    lag1 = float(np.corrcoef(left, right)[0, 1]) if np.std(left) > 0 and np.std(right) > 0 else np.nan
    return {
        "lag1_residual_autocorrelation": lag1,
        "durbin_watson": float(np.sum(np.diff(residual) ** 2) / denominator),
        "material_residual_autocorrelation": bool(np.isfinite(lag1) and abs(lag1) >= 0.3),
    }


def _annualized_endpoint_change(years: np.ndarray, values: np.ndarray) -> float:
    span = int(years[-1] - years[0])
    if span <= 0 or values[0] <= 0 or values[-1] <= 0:
        return np.nan
    return 100.0 * (math.exp(math.log(values[-1] / values[0]) / span) - 1.0)


def segmented_summary(panel: pd.DataFrame) -> tuple[dict, pd.DataFrame, pd.DataFrame]:
    """Fit a descriptive segmented curve to annual posterior means."""
    panel = panel.sort_values("year").copy()
    years = panel.year.to_numpy(int)
    values = panel.val.to_numpy(float)
    y = np.log(values)
    selected = select_segmented_bic(years, y)

    contrast = np.zeros(len(selected["beta"]))
    contrast[1] = 1.0
    for j, knot in enumerate(selected["knots"]):
        contrast[2 + j] = (years.max() - knot) / (years.max() - years.min())
    slope = float(contrast @ selected["beta"])
    aapc = 100.0 * (math.exp(slope) - 1.0)
    diagnostics = residual_diagnostics(selected["residual"])
    summary = {
        "location_name": panel.location_name.iloc[0], "sex_name": panel.sex_name.iloc[0],
        "measure_name": panel.measure_name.iloc[0], "start_year": int(years.min()), "end_year": int(years.max()),
        "joinpoint_count": len(selected["knots"]), "joinpoint_years": ",".join(map(str, selected["knots"])),
        "selection_method": "minimum BIC; breakpoint location counted in penalty; descriptive only",
        "selected_bic": selected["bic"],
        "candidate_bic_0_joinpoints": selected["candidate_bic"][0],
        "candidate_bic_1_joinpoint": selected["candidate_bic"][1],
        "candidate_bic_2_joinpoints": selected["candidate_bic"][2],
        "aapc": aapc,
        "observed_annualized_endpoint_change_pct": _annualized_endpoint_change(years, values),
        "formal_inference_performed": False,
        "model_label": "descriptive BIC-selected segmented log-linear curve; not Joinpoint",
        "inference_note": DESCRIPTIVE_INFERENCE_NOTE,
        **diagnostics,
    }
    segments = []
    breaks = (int(years.min()), *selected["knots"], int(years.max()))
    for index, (start, end) in enumerate(zip(breaks[:-1], breaks[1:])):
        c = np.zeros(len(selected["beta"]))
        c[1] = 1.0
        for j, knot in enumerate(selected["knots"]):
            c[2 + j] = 1.0 if start >= knot else 0.0
        seg_slope = float(c @ selected["beta"])
        segments.append({
            "location_name": summary["location_name"], "sex_name": summary["sex_name"], "measure_name": summary["measure_name"],
            "segment_index": index + 1, "start_year": start, "end_year": end,
            "apc": 100.0 * (math.exp(seg_slope) - 1.0),
            "formal_inference_performed": False, "inference_note": DESCRIPTIVE_INFERENCE_NOTE,
        })
    fitted = panel[["location_name", "sex_name", "measure_name", "year", "val", "lower", "upper"]].copy()
    fitted["fitted"] = np.exp(selected["fitted"])
    fitted["joinpoint_years"] = summary["joinpoint_years"]
    return summary, pd.DataFrame(segments), fitted


def segmented_specification_sensitivity(df: pd.DataFrame) -> pd.DataFrame:
    """Assess qualitative stability across reasonable descriptive curve choices."""
    specifications = (
        ("primary", 2, 4, 2023, "log_rate"),
        ("max_1_breakpoint", 1, 4, 2023, "log_rate"),
        ("max_3_breakpoints", 3, 4, 2023, "log_rate"),
        ("minimum_5_year_segments", 2, 5, 2023, "log_rate"),
        ("prepandemic_1990_2019", 2, 4, 2019, "log_rate"),
        ("linear_rate_scale", 2, 4, 2023, "rate"),
    )
    trend = df[
        df.measure_name.isin(OUTCOMES)
        & df.metric_name.eq("Rate")
        & df.age_name.eq(ASR)
    ]
    rows = []
    for (location, sex, outcome), full_panel in trend.groupby(
        ["location_name", "sex_name", "measure_name"], sort=True
    ):
        primary_direction = None
        group_rows = []
        for specification, max_joinpoints, min_segment, end_year, scale in specifications:
            panel = full_panel[full_panel.year.le(end_year)].sort_values("year")
            years = panel.year.to_numpy(int)
            values = panel.val.to_numpy(float)
            response = np.log(values) if scale == "log_rate" else values
            selected = select_segmented_bic(
                years,
                response,
                max_joinpoints=max_joinpoints,
                min_segment=min_segment,
            )
            fitted_values = (
                np.exp(selected["fitted"])
                if scale == "log_rate"
                else selected["fitted"]
            )
            annualized = _annualized_endpoint_change(years, fitted_values)
            direction = trend_direction(annualized)
            if specification == "primary":
                primary_direction = direction
            group_rows.append(
                {
                    "location_name": location,
                    "sex_name": sex,
                    "measure_name": outcome,
                    "specification": specification,
                    "model_scale": scale,
                    "start_year": int(years.min()),
                    "end_year": int(years.max()),
                    "maximum_breakpoints": max_joinpoints,
                    "minimum_segment_years": min_segment,
                    "selected_breakpoint_count": len(selected["knots"]),
                    "selected_breakpoint_years": ",".join(
                        map(str, selected["knots"])
                    ),
                    "fitted_annualized_endpoint_change_pct": annualized,
                    "trajectory_direction": direction,
                    "formal_inference_performed": False,
                    "inference_note": DESCRIPTIVE_INFERENCE_NOTE,
                }
            )
        for row in group_rows:
            row["primary_direction"] = primary_direction
            row["direction_stable_vs_primary"] = (
                row["trajectory_direction"] == primary_direction
            )
            rows.append(row)
    return pd.DataFrame(rows)


def run_segmented(df: pd.DataFrame, end_year: int = 2023):
    """Run deterministic descriptive trend summaries."""
    trend = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR) & (df.year <= end_year)]
    summaries, segments, fitted = [], [], []
    for _, panel in trend.groupby(["location_name", "sex_name", "measure_name"], sort=True):
        a, b, c = segmented_summary(panel)
        summaries.append(a)
        segments.append(b)
        fitted.append(c)
    return pd.DataFrame(summaries), pd.concat(segments, ignore_index=True), pd.concat(fitted, ignore_index=True)


def weighted_trend_sensitivity(df: pd.DataFrame, primary: pd.DataFrame) -> pd.DataFrame:
    trend = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR)]
    rows = []
    for _, row in primary.iterrows():
        p = trend[(trend.location_name == row.location_name) & (trend.sex_name == row.sex_name) & (trend.measure_name == row.measure_name)].sort_values("year")
        years = p.year.to_numpy(int)
        y = np.log(p.val.to_numpy(float))
        se = (np.log(p.upper.to_numpy(float)) - np.log(p.lower.to_numpy(float))) / (2 * 1.96)
        weights = 1.0 / np.maximum(se**2, 1e-12)
        knots = tuple(int(x) for x in str(row.joinpoint_years).split(",") if str(x).strip())
        fit = fit_linear(y, segmented_design(years, knots), weights)
        aapc = 100.0 * (math.exp((fit["fitted"][-1] - fit["fitted"][0]) / (years[-1] - years[0])) - 1.0)
        rows.append({"location_name": row.location_name, "sex_name": row.sex_name, "measure_name": row.measure_name,
                     "primary_aapc": row.aapc, "ui_weighted_fixed_knot_aapc": aapc,
                     "difference": aapc - row.aapc, "interpretation": "Sensitivity only; GBD UIs are not independent sampling SEs."})
    return pd.DataFrame(rows)


def trend_direction(
    value: float, tolerance: float = PRACTICAL_STABILITY_THRESHOLD_PCT_PER_YEAR
) -> str:
    """Return a descriptive trajectory label using the practical-stability band."""
    if not np.isfinite(value):
        return "not available"
    if value > tolerance:
        return "increase"
    if value < -tolerance:
        return "decrease"
    return "practically stable"


def practical_stability_sensitivity(segmented: pd.DataFrame) -> pd.DataFrame:
    """Show how compact trajectory labels change across descriptive thresholds."""
    rows = []
    for row in segmented.itertuples(index=False):
        item = {
            "location_name": row.location_name,
            "sex_name": row.sex_name,
            "measure_name": row.measure_name,
            "aapc_pct_per_year": float(row.aapc),
        }
        for threshold in PRACTICAL_STABILITY_SENSITIVITY_THRESHOLDS:
            item[f"label_at_{threshold:.2f}_pct_per_year"] = trend_direction(
                float(row.aapc), tolerance=threshold
            )
        rows.append(item)
    out = pd.DataFrame(rows)
    out["primary_threshold_pct_per_year"] = PRACTICAL_STABILITY_THRESHOLD_PCT_PER_YEAR
    out["interpretation"] = (
        "Practical-description sensitivity only; labels are not statistical equivalence tests."
    )
    return out


def descriptive_trajectory_contrast(a: pd.DataFrame, b: pd.DataFrame) -> dict:
    """Compare annual point-estimate trajectories without inferential claims."""
    a = a.sort_values("year")
    b = b.sort_values("year")
    years = a.year.to_numpy(int)
    if not np.array_equal(years, b.year.to_numpy(int)):
        raise ValueError("Trajectory contrasts require identical years.")
    values_a, values_b = a.val.to_numpy(float), b.val.to_numpy(float)
    changes_a = 100.0 * np.diff(np.log(values_a))
    changes_b = 100.0 * np.diff(np.log(values_b))
    correlation = (
        float(np.corrcoef(changes_a, changes_b)[0, 1])
        if np.std(changes_a) > 0 and np.std(changes_b) > 0
        else np.nan
    )
    endpoint_a = _annualized_endpoint_change(years, values_a)
    endpoint_b = _annualized_endpoint_change(years, values_b)
    return {
        "group_a_annualized_endpoint_change_pct": endpoint_a,
        "group_b_annualized_endpoint_change_pct": endpoint_b,
        "annualized_endpoint_change_difference_b_minus_a_pct_points": endpoint_b - endpoint_a,
        "rms_annual_log_change_difference": float(np.sqrt(np.mean((changes_b - changes_a) ** 2))),
        "annual_log_change_correlation": correlation,
        "same_annualized_endpoint_change_direction": trend_direction(endpoint_a) == trend_direction(endpoint_b),
    }


def build_trajectory_contrasts(df: pd.DataFrame) -> pd.DataFrame:
    """Create descriptive country and sex trajectory contrasts."""
    trend = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR)]
    rows = []
    for family in ("country", "sex"):
        if family == "country":
            specs = [(sex, outcome) for sex in SEXES for outcome in OUTCOMES]
            for sex, outcome in specs:
                a = trend[(trend.location_name == LOCATIONS[0]) & (trend.sex_name == sex) & (trend.measure_name == outcome)]
                b = trend[(trend.location_name == LOCATIONS[1]) & (trend.sex_name == sex) & (trend.measure_name == outcome)]
                contrast = descriptive_trajectory_contrast(a, b)
                rows.append({"comparison_family": family, "stratum": sex, "measure_name": outcome,
                             "group_a": LOCATIONS[0], "group_b": LOCATIONS[1], **contrast})
        else:
            specs = [(loc, outcome) for loc in LOCATIONS for outcome in OUTCOMES]
            for loc, outcome in specs:
                a = trend[(trend.location_name == loc) & (trend.sex_name == SEXES[0]) & (trend.measure_name == outcome)]
                b = trend[(trend.location_name == loc) & (trend.sex_name == SEXES[1]) & (trend.measure_name == outcome)]
                contrast = descriptive_trajectory_contrast(a, b)
                rows.append({"comparison_family": family, "stratum": loc, "measure_name": outcome,
                             "group_a": SEXES[0], "group_b": SEXES[1], **contrast})
    out = pd.DataFrame(rows)
    out["formal_inference_performed"] = False
    out["method_note"] = DESCRIPTIVE_INFERENCE_NOTE
    return out


def decompose_change(pop0: np.ndarray, pop1: np.ndarray, rate0: np.ndarray, rate1: np.ndarray) -> dict:
    n0, n1 = pop0.sum(), pop1.sum()
    s0, s1 = pop0 / n0, pop1 / n1
    base = {"N": n0, "S": s0, "R": rate0}
    end = {"N": n1, "S": s1, "R": rate1}
    names = {
        "N": "population_size_change",
        "S": "age_structure_change",
        "R": "age_specific_rate_change",
    }
    out = dict.fromkeys(names.values(), 0.0)
    def total(x): return float(x["N"] * np.sum(x["S"] * x["R"]))
    for order in permutations(("N", "S", "R")):
        cur = {k: np.array(v, copy=True) if isinstance(v, np.ndarray) else v for k, v in base.items()}
        for factor in order:
            before = total(cur)
            cur[factor] = (
                np.array(end[factor], copy=True)
                if isinstance(end[factor], np.ndarray)
                else end[factor]
            )
            out[names[factor]] += (total(cur) - before) / 6.0
    out["total_change"] = total(end) - total(base)
    out["component_sum"] = sum(out[name] for name in names.values())
    out["closure_error"] = out["component_sum"] - out["total_change"]
    return out


def run_decomposition(
    df: pd.DataFrame,
    pop: pd.DataFrame,
    windows=((1990, 2023), (2000, 2023), (2010, 2023)),
    ages: tuple[str, ...] | None = None,
    outcomes: tuple[str, ...] = OUTCOMES,
) -> pd.DataFrame:
    ages = ages or select_decomposition_ages(set(pop.age_name.dropna().astype(str)))
    rates = df[(df.measure_name.isin(outcomes)) & (df.metric_name == "Rate") & df.age_name.isin(ages)][
        ["location_name", "sex_name", "measure_name", "age_name", "year", "val"]
    ].rename(columns={"val": "rate"})
    merged = rates.merge(pop, on=["location_name", "sex_name", "age_name", "year"], validate="many_to_one")
    rows = []
    for loc in LOCATIONS:
        for sex in SEXES:
            for outcome in outcomes:
                p = merged[(merged.location_name == loc) & (merged.sex_name == sex) & (merged.measure_name == outcome)]
                for start, end in windows:
                    a = p[p.year == start].set_index("age_name").reindex(ages)
                    b = p[p.year == end].set_index("age_name").reindex(ages)
                    if a[["population", "rate"]].isna().any().any() or b[["population", "rate"]].isna().any().any():
                        raise ValueError(
                            f"Decomposition is missing age-specific cells for {loc}, {sex}, "
                            f"{outcome}, {start}-{end}."
                        )
                    result = decompose_change(a.population.to_numpy(), b.population.to_numpy(), a.rate.to_numpy() / 100000.0, b.rate.to_numpy() / 100000.0)
                    base_count = float(np.sum(a.population * a.rate / 100000.0))
                    end_count = float(np.sum(b.population * b.rate / 100000.0))
                    row = {"location_name": loc, "sex_name": sex, "measure_name": outcome,
                           "start_year": start, "end_year": end, "all_age_count_start_reconstructed": base_count,
                           "all_age_count_end_reconstructed": end_count, "population_source": pop.population_source.iloc[0],
                           "age_partition": (
                               "fine_5_year" if ages == FINE_DECOMPOSITION_AGES
                               else "supported_10_79" if ages == tuple(FINE_DECOMPOSITION_AGES[2:16])
                               else "custom_or_provisional"
                           ),
                           "age_group_count": len(ages), **result}
                    for component in COMPONENT_COLORS:
                        row[f"{component}_pct_of_total"] = 100.0 * row[component] / row["total_change"] if row["total_change"] else np.nan
                    rows.append(row)
    return pd.DataFrame(rows)


def incidence_supported_age_decomposition(
    df: pd.DataFrame, pop: pd.DataFrame, start_year: int = 1990, end_year: int = 2023
) -> pd.DataFrame:
    """Sensitivity restricted to incidence ages with positive source-export rates."""
    ages = tuple(FINE_DECOMPOSITION_AGES[2:16])  # 10-14 through 75-79
    burden_ages = set(df.loc[df.measure_name.eq("Incidence"), "age_name"].astype(str))
    population_ages = set(pop.age_name.astype(str))
    if not set(ages).issubset(burden_ages & population_ages):
        return pd.DataFrame(columns=[
            "location_name", "sex_name", "measure_name", "start_year", "end_year",
            "supported_age_quantity_start_reconstructed",
            "supported_age_quantity_end_reconstructed", "population_source",
            "age_partition", "quantity_scope", "age_group_count",
            "population_size_change", "age_structure_change",
            "age_specific_rate_change", "total_change", "component_sum",
            "closure_error", "population_size_change_pct_of_total",
            "age_structure_change_pct_of_total", "age_specific_rate_change_pct_of_total",
        ])
    result = run_decomposition(
        df,
        pop,
        windows=((start_year, end_year),),
        ages=ages,
        outcomes=("Incidence",),
    ).rename(
        columns={
            "all_age_count_start_reconstructed": "supported_age_quantity_start_reconstructed",
            "all_age_count_end_reconstructed": "supported_age_quantity_end_reconstructed",
        }
    )
    result.insert(
        result.columns.get_loc("age_partition") + 1,
        "quantity_scope",
        "incidence ages 10-79; sensitivity only; not all-age burden",
    )
    return result


def _collapsed_age_group(age_name: str) -> str:
    match = re.match(r"(\d+)", str(age_name))
    if not match:
        raise ValueError(f"Cannot collapse unrecognized age label {age_name!r}.")
    lower = int(match.group(1))
    if lower < 20:
        return "0-19 years"
    if lower < 40:
        return "20-39 years"
    if lower < 60:
        return "40-59 years"
    return "60+ years"


def decomposition_age_bin_sensitivity(
    df: pd.DataFrame, pop: pd.DataFrame, start_year: int = 1990, end_year: int = 2023
) -> pd.DataFrame:
    """Compare the finest available decomposition with four collapsed age groups."""
    ages = select_decomposition_ages(set(pop.age_name.dropna().astype(str)))
    baseline = run_decomposition(df, pop, windows=((start_year, end_year),))
    rates = df[
        df.measure_name.isin(OUTCOMES)
        & df.metric_name.eq("Rate")
        & df.age_name.isin(ages)
        & df.year.isin((start_year, end_year))
    ][["location_name", "sex_name", "measure_name", "age_name", "year", "val"]].rename(
        columns={"val": "rate"}
    )
    merged = rates.merge(
        pop[["location_name", "sex_name", "age_name", "year", "population"]],
        on=["location_name", "sex_name", "age_name", "year"],
        validate="many_to_one",
    )
    merged["collapsed_age_group"] = merged.age_name.map(_collapsed_age_group)
    merged["count"] = merged.population * merged.rate / 100_000.0
    collapsed = merged.groupby(
        ["location_name", "sex_name", "measure_name", "year", "collapsed_age_group"],
        as_index=False,
        observed=True,
    ).agg(population=("population", "sum"), count=("count", "sum"))
    collapsed["rate"] = collapsed["count"] / collapsed.population

    component_names = tuple(COMPONENT_COLORS)
    rows = []
    for (location, sex, outcome), panel in collapsed.groupby(
        ["location_name", "sex_name", "measure_name"], sort=True
    ):
        age_order = ("0-19 years", "20-39 years", "40-59 years", "60+ years")
        first = panel[panel.year.eq(start_year)].set_index("collapsed_age_group").reindex(age_order)
        last = panel[panel.year.eq(end_year)].set_index("collapsed_age_group").reindex(age_order)
        if first[["population", "rate"]].isna().any().any() or last[["population", "rate"]].isna().any().any():
            raise ValueError("Collapsed decomposition sensitivity has missing age-year cells.")
        coarse = decompose_change(
            first.population.to_numpy(),
            last.population.to_numpy(),
            first.rate.to_numpy(),
            last.rate.to_numpy(),
        )
        fine = baseline[
            baseline.location_name.eq(location)
            & baseline.sex_name.eq(sex)
            & baseline.measure_name.eq(outcome)
        ].iloc[0]
        fine_rank = tuple(
            sorted(component_names, key=lambda name: abs(float(fine[name])), reverse=True)
        )
        coarse_rank = tuple(
            sorted(component_names, key=lambda name: abs(float(coarse[name])), reverse=True)
        )
        total_scale = max(abs(float(fine.total_change)), np.finfo(float).eps)
        shifts = {
            name: abs(float(fine[name]) - float(coarse[name])) / total_scale * 100.0
            for name in component_names
        }
        sign_stability = {
            name: bool(np.sign(float(fine[name])) == np.sign(float(coarse[name])))
            for name in component_names
        }
        row = {
            "location_name": location,
            "sex_name": sex,
            "measure_name": outcome,
            "start_year": start_year,
            "end_year": end_year,
            "finest_age_partition": (
                "fine_5_year" if ages == FINE_DECOMPOSITION_AGES else "provisional_broad_end_groups"
            ),
            "finest_age_group_count": len(ages),
            "collapsed_age_group_count": len(age_order),
            "finest_component_rank": " > ".join(fine_rank),
            "collapsed_component_rank": " > ".join(coarse_rank),
            "component_rank_stable": fine_rank == coarse_rank,
            "maximum_component_shift_pct_of_total_change": max(shifts.values()),
        }
        for name in component_names:
            row[f"finest_{name}"] = float(fine[name])
            row[f"collapsed_{name}"] = float(coarse[name])
            row[f"{name}_sign_stable"] = sign_stability[name]
            row[f"{name}_shift_pct_of_total_change"] = shifts[name]
        row["material_age_bin_sensitivity"] = bool(
            not all(sign_stability.values())
            or fine_rank != coarse_rank
            or max(shifts.values()) >= 10.0
        )
        row["sensitivity_rule"] = (
            "Flagged if a component changes sign, absolute-magnitude rank changes, "
            "or any component shifts by at least 10% of total count change."
        )
        rows.append(row)
    return pd.DataFrame(rows)


def chained_decomposition(df: pd.DataFrame, pop: pd.DataFrame, step: int) -> pd.DataFrame:
    years = list(range(1990, 2024, step))
    if years[-1] != 2023: years.append(2023)
    pieces = []
    for start, end in zip(years[:-1], years[1:]):
        pieces.append(run_decomposition(df, pop, windows=((start, end),)))
    out = pd.concat(pieces, ignore_index=True).sort_values(["location_name", "sex_name", "measure_name", "end_year"])
    for col in (*COMPONENT_COLORS, "total_change"):
        out[f"cumulative_{col}"] = out.groupby(["location_name", "sex_name", "measure_name"])[col].cumsum()
    out["chain_step_years"] = step
    return out


def decomposition_path_sensitivity_summary(
    endpoint: pd.DataFrame, annual: pd.DataFrame, fiveyear: pd.DataFrame
) -> pd.DataFrame:
    """Compare endpoint Shapley allocations with chained allocation paths."""
    keys = ["location_name", "sex_name", "measure_name"]
    components = tuple(COMPONENT_COLORS)
    baseline = endpoint[endpoint.start_year.eq(1990) & endpoint.end_year.eq(2023)][
        keys + ["total_change", *components]
    ].copy()
    out = baseline
    for label, chained in (("annual", annual), ("fiveyear", fiveyear)):
        last = chained.sort_values("end_year").groupby(keys, as_index=False).tail(1)
        columns = {
            f"cumulative_{component}": f"{label}_chained_{component}"
            for component in components
        }
        selected = last[keys + list(columns)].rename(columns=columns)
        out = out.merge(selected, on=keys, validate="one_to_one")
        for component in components:
            out[f"{label}_{component}_shift_pct_of_total_change"] = (
                100.0
                * (out[f"{label}_chained_{component}"] - out[component]).abs()
                / out.total_change.abs().clip(lower=np.finfo(float).eps)
            )
    shift_columns = [column for column in out if column.endswith("shift_pct_of_total_change")]
    out["maximum_path_shift_pct_of_total_change"] = out[shift_columns].max(axis=1)
    out["interpretation"] = (
        "Difference in accounting allocation under endpoint, annual-chain, and five-year-chain "
        "replacement paths; not statistical uncertainty."
    )
    return out


def apc_period(year: int) -> str | None:
    """Compatibility wrapper for the primary six-period APC window."""
    return apc.PRIMARY_WINDOW.period_for_year(year)


def run_secondary_apc(
    df: pd.DataFrame,
    pop: pd.DataFrame,
    include_last_period: bool = True,
    weighting: str = "population",
) -> dict[str, pd.DataFrame]:
    """Run the primary 1994--2023 or sensitivity 1990--2019 APC model."""
    window = apc.PRIMARY_WINDOW if include_last_period else apc.SENSITIVITY_WINDOW
    return apc.run_apc(
        df, pop, window, LOCATIONS, SEXES, measures=OUTCOMES, weighting=weighting
    )


def compare_apc_weighting(
    weighted: dict[str, pd.DataFrame], unweighted: dict[str, pd.DataFrame]
) -> pd.DataFrame:
    """Summarize the sensitivity of APC point estimates to population weighting."""
    keys = ["location_name", "sex_name", "measure_name"]
    left = weighted["summary"][keys + ["net_drift"]].rename(
        columns={"net_drift": "population_weighted_global_period_slope"}
    )
    right = unweighted["summary"][keys + ["net_drift"]].rename(
        columns={"net_drift": "equal_weight_global_period_slope"}
    )
    out = left.merge(right, on=keys, validate="one_to_one")
    out["equal_minus_population_weighted_global_period_slope"] = (
        out.equal_weight_global_period_slope - out.population_weighted_global_period_slope
    )
    curve_specs = (
        ("local_drift", "age_name", "local_drift", "maximum_absolute_age_specific_slope_difference"),
        ("age_curve", "age_name", "longitudinal_age_rr", "maximum_absolute_log_age_rate_ratio_difference"),
        ("period_rr", "period", "period_rr", "maximum_absolute_log_period_rate_ratio_difference"),
        ("cohort_rr", "cohort_midpoint", "cohort_rr", "maximum_absolute_log_cohort_rate_ratio_difference"),
    )
    for table, coordinate, value, output_name in curve_specs:
        joined = weighted[table][keys + [coordinate, value]].merge(
            unweighted[table][keys + [coordinate, value]],
            on=keys + [coordinate],
            suffixes=("_population", "_equal"),
            validate="one_to_one",
        )
        if value == "local_drift":
            joined["difference"] = (
                joined[f"{value}_equal"] - joined[f"{value}_population"]
            ).abs()
        else:
            joined["difference"] = np.abs(
                np.log(joined[f"{value}_equal"] / joined[f"{value}_population"])
            )
        maximum = joined.groupby(keys, as_index=False).difference.max().rename(
            columns={"difference": output_name}
        )
        out = out.merge(maximum, on=keys, validate="one_to_one")
    out["interpretation"] = (
        "Descriptive weighting sensitivity; differences are point-estimate changes and "
        "are not GBD posterior uncertainty."
    )
    return out


def format_apc_tables(
    results: dict[str, pd.DataFrame], prefix: str
) -> dict[str, pd.DataFrame]:
    """Relabel custom APC outputs so they are not confused with NCI estimable functions."""
    method = (
        "Custom descriptive weighted least-squares age-period-cohort summaries; "
        "not asserted equivalent to conventional NCI APC estimable functions."
    )
    role = np.where(
        results["summary"].measure_name.eq("Incidence"),
        "principal secondary APC outcome",
        "exploratory rate-surface extension",
    )
    summary = results["summary"].rename(
        columns={
            "net_drift": "global_period_slope_pct_per_year",
            "weighted_log_rate_rss": "log_rate_objective_value",
        }
    ).copy()
    summary["analysis_role"] = role
    summary["method_label"] = method
    local = results["local_drift"].rename(
        columns={"local_drift": "age_specific_slope_pct_per_year"}
    ).copy()
    local["method_label"] = method
    age = results["age_curve"].rename(
        columns={"longitudinal_age_rr": "descriptive_age_rate_ratio"}
    ).copy()
    age["method_label"] = method
    period = results["period_rr"].rename(
        columns={"period_rr": "period_curvature_rate_ratio"}
    ).copy()
    period["method_label"] = method
    cohort = results["cohort_rr"].rename(
        columns={"cohort_rr": "cohort_curvature_rate_ratio"}
    ).copy()
    cohort["method_label"] = method
    cells = results["cells"].rename(
        columns={"rate": "observed_rate", "fitted_rate": "custom_fitted_rate"}
    ).copy()
    cells["method_label"] = method
    return {
        f"{prefix}_summary": summary,
        f"{prefix}_age_specific_slopes": local,
        f"{prefix}_age_curve": age,
        f"{prefix}_period_curvature": period,
        f"{prefix}_cohort_curvature": cohort,
        f"{prefix}_cells": cells,
    }


def compare_primary_apc_directions(
    df: pd.DataFrame,
    primary: pd.DataFrame,
    apc_summary: pd.DataFrame,
) -> pd.DataFrame:
    """Compare incidence trends with the custom APC global period slope."""
    incidence = df[
        (df.measure_name == "Incidence")
        & (df.metric_name == "Rate")
        & (df.age_name == ASR)
    ]
    endpoint_rows = []
    for (loc, sex), panel in incidence.groupby(["location_name", "sex_name"], sort=True):
        panel = panel.sort_values("year")
        endpoint_rows.append({
            "location_name": loc,
            "sex_name": sex,
            "primary_observed_annualized_endpoint_change_pct": _annualized_endpoint_change(
                panel.year.to_numpy(int), panel.val.to_numpy(float)
            ),
        })
    endpoint = pd.DataFrame(endpoint_rows)
    primary_incidence = primary[primary.measure_name == "Incidence"][
        ["location_name", "sex_name", "aapc"]
    ].rename(columns={"aapc": "primary_segmented_aapc"})
    apc_incidence = apc_summary[apc_summary.measure_name == "Incidence"][
        ["location_name", "sex_name", "net_drift"]
    ].rename(columns={"net_drift": "apc_global_period_slope"})
    out = primary_incidence.merge(endpoint, on=["location_name", "sex_name"], validate="one_to_one")
    out = out.merge(apc_incidence, on=["location_name", "sex_name"], validate="one_to_one")
    out.insert(2, "measure_name", "Incidence")
    out["primary_segmented_direction"] = out.primary_segmented_aapc.map(trend_direction)
    out["primary_observed_direction"] = out.primary_observed_annualized_endpoint_change_pct.map(trend_direction)
    out["apc_global_period_slope_label"] = out.apc_global_period_slope.map(trend_direction)
    out["apc_vs_segmented_direction_agreement"] = (
        out.apc_global_period_slope_label == out.primary_segmented_direction
    )
    out["apc_vs_observed_direction_agreement"] = (
        out.apc_global_period_slope_label == out.primary_observed_direction
    )
    out["apc_minus_segmented_aapc_pct_points"] = out.apc_global_period_slope - out.primary_segmented_aapc
    out["apc_minus_observed_annualized_endpoint_change_pct_points"] = (
        out.apc_global_period_slope - out.primary_observed_annualized_endpoint_change_pct
    )
    out["comparison_note"] = (
        "Labels apply the 0.05%/year practical-stability band to descriptive point estimates; "
        "agreement is not a significance or equivalence test."
    )
    return out


def compare_apc_windows(
    primary_summary: pd.DataFrame,
    sensitivity_summary: pd.DataFrame,
) -> pd.DataFrame:
    """Compare custom global period slopes across the two complete windows."""
    keys = ["location_name", "sex_name", "measure_name"]
    primary = primary_summary[keys + ["net_drift"]].rename(
        columns={"net_drift": "primary_global_period_slope_1994_2023"}
    )
    sensitivity = sensitivity_summary[keys + ["net_drift"]].rename(
        columns={"net_drift": "sensitivity_global_period_slope_1990_2019"}
    )
    out = primary.merge(sensitivity, on=keys, validate="one_to_one")
    out["primary_direction"] = out.primary_global_period_slope_1994_2023.map(trend_direction)
    out["sensitivity_direction"] = out.sensitivity_global_period_slope_1990_2019.map(
        trend_direction
    )
    out["direction_agreement"] = out.primary_direction.eq(out.sensitivity_direction)
    out["global_period_slope_difference_pct_points"] = (
        out.primary_global_period_slope_1994_2023 - out.sensitivity_global_period_slope_1990_2019
    )
    out["comparison_note"] = (
        "Labels compare descriptive point estimates using the practical-stability band across "
        "two six-period windows; this is not an inferential or equivalence test."
    )
    return out


def _extreme_pattern(
    frame: pd.DataFrame, value_column: str, label_column: str, digits: int = 3
) -> str:
    if frame.empty:
        return "not analyzed"
    low = frame.loc[frame[value_column].idxmin()]
    high = frame.loc[frame[value_column].idxmax()]
    return (
        f"minimum {low[value_column]:.{digits}f} at {low[label_column]}; "
        f"maximum {high[value_column]:.{digits}f} at {high[label_column]}"
    )


def build_cross_analysis_consistency(
    burden: pd.DataFrame,
    population: pd.DataFrame,
    endpoints: pd.DataFrame,
    segmented: pd.DataFrame,
    apc_results: dict[str, pd.DataFrame],
    decomposition: pd.DataFrame,
) -> pd.DataFrame:
    """Place complementary estimands side by side without treating them as replications."""
    rows = []
    for location in LOCATIONS:
        for sex in SEXES:
            for outcome in OUTCOMES:
                endpoint = endpoints[
                    endpoints.location_name.eq(location)
                    & endpoints.sex_name.eq(sex)
                    & endpoints.measure_name.eq(outcome)
                    & endpoints.metric_name.eq("Age-standardized rate per 100,000")
                ].iloc[0]
                count_endpoint = endpoints[
                    endpoints.location_name.eq(location)
                    & endpoints.sex_name.eq(sex)
                    & endpoints.measure_name.eq(outcome)
                    & endpoints.metric_name.eq("All-age count")
                ].iloc[0]
                trend = segmented[
                    segmented.location_name.eq(location)
                    & segmented.sex_name.eq(sex)
                    & segmented.measure_name.eq(outcome)
                ].iloc[0]
                decomp = decomposition[
                    decomposition.location_name.eq(location)
                    & decomposition.sex_name.eq(sex)
                    & decomposition.measure_name.eq(outcome)
                    & decomposition.start_year.eq(1990)
                    & decomposition.end_year.eq(2023)
                ].iloc[0]
                row = {
                    "location_name": location,
                    "sex_name": sex,
                    "measure_name": outcome,
                    "asr_endpoint_percent_change_1990_2023": float(
                        endpoint.percent_change_point_estimate
                    ),
                    "all_age_count_percent_change_1990_2023": float(
                        count_endpoint.percent_change_point_estimate
                    ),
                    "segmented_aapc_1990_2023": float(trend.aapc),
                    "segmented_overall_direction": trend_direction(float(trend.aapc)),
                    "segmented_breakpoint_years": trend.joinpoint_years,
                    "population_size_change": float(decomp.population_size_change),
                    "age_structure_change": float(decomp.age_structure_change),
                    "age_specific_rate_change": float(decomp.age_specific_rate_change),
                    "decomposition_total_change": float(decomp.total_change),
                    "decomposition_age_partition": decomp.age_partition,
                    "apc_global_period_slope_1994_2023": np.nan,
                    "apc_global_period_slope_label": "not analyzed",
                    "important_age_specific_slopes": "not analyzed",
                    "age_specific_slopes_include_opposing_directions": False,
                    "period_curvature_pattern": "not analyzed",
                    "cohort_curvature_pattern": "not analyzed",
                    "methods_are_independent_replications": False,
                    "uncertainty_note": DESCRIPTIVE_INFERENCE_NOTE,
                }
                panel_filter = (
                    apc_results["summary"].location_name.eq(location)
                    & apc_results["summary"].sex_name.eq(sex)
                    & apc_results["summary"].measure_name.eq(outcome)
                )
                apc_row = apc_results["summary"][panel_filter].iloc[0]
                local = apc_results["local_drift"]
                local = local[
                    local.location_name.eq(location)
                    & local.sex_name.eq(sex)
                    & local.measure_name.eq(outcome)
                ]
                period = apc_results["period_rr"]
                period = period[
                    period.location_name.eq(location)
                    & period.sex_name.eq(sex)
                    & period.measure_name.eq(outcome)
                ]
                cohort = apc_results["cohort_rr"]
                cohort = cohort[
                    cohort.location_name.eq(location)
                    & cohort.sex_name.eq(sex)
                    & cohort.measure_name.eq(outcome)
                ]
                row.update(
                    {
                        "apc_global_period_slope_1994_2023": float(apc_row.net_drift),
                        "apc_global_period_slope_label": trend_direction(float(apc_row.net_drift)),
                        "important_age_specific_slopes": _extreme_pattern(
                            local, "local_drift", "age_name"
                        ),
                        "age_specific_slopes_include_opposing_directions": bool(
                            local.local_drift.min() < 0 < local.local_drift.max()
                        ),
                        "period_curvature_pattern": _extreme_pattern(period, "period_rr", "period"),
                        "cohort_curvature_pattern": _extreme_pattern(
                            cohort, "cohort_rr", "cohort_midpoint"
                        ),
                    }
                )
                rows.append(row)
    return pd.DataFrame(rows)


def investigate_cross_method_contradictions(
    burden: pd.DataFrame,
    population: pd.DataFrame,
    consistency: pd.DataFrame,
) -> pd.DataFrame:
    """Diagnose apparent cross-method contradictions using aligned contrasts."""
    rows = []
    ages = apc.select_apc_ages(set(burden.age_name.dropna().astype(str)))
    for _, item in consistency.iterrows():
        location, sex, outcome = item.location_name, item.sex_name, item.measure_name
        asr = burden[
            burden.location_name.eq(location)
            & burden.sex_name.eq(sex)
            & burden.measure_name.eq(outcome)
            & burden.metric_name.eq("Rate")
            & burden.age_name.eq(ASR)
            & burden.year.between(1994, 2023)
        ].sort_values("year")
        asr_aligned = _annualized_endpoint_change(
            asr.year.to_numpy(int), asr.val.to_numpy(float)
        )
        keys = ["location_name", "sex_name", "age_name", "year"]
        age_rates = burden[
            burden.location_name.eq(location)
            & burden.sex_name.eq(sex)
            & burden.measure_name.eq(outcome)
            & burden.metric_name.eq("Rate")
            & burden.age_name.isin(ages)
            & burden.year.isin((1994, 2023))
        ][keys + ["val"]].rename(columns={"val": "rate"})
        age_population = population[
            population.location_name.eq(location)
            & population.sex_name.eq(sex)
            & population.age_name.isin(ages)
            & population.year.isin((1994, 2023))
        ][keys + ["population"]]
        crude = age_rates.merge(age_population, on=keys, validate="one_to_one")
        crude["events"] = crude.population * crude.rate / 100_000.0
        crude = crude.groupby("year", as_index=False).agg(
            events=("events", "sum"), population=("population", "sum")
        )
        crude["rate"] = crude.events / crude.population * 100_000.0
        crude_aligned = _annualized_endpoint_change(
            crude.year.to_numpy(int), crude.rate.to_numpy(float)
        )
        directions = {
            "segmented_1990_2023": item.segmented_overall_direction,
            "asr_endpoint_1994_2023": trend_direction(asr_aligned),
            "selected_age_crude_1994_2023": trend_direction(crude_aligned),
            "apc_global_period_slope_1994_2023": item.apc_global_period_slope_label,
        }
        disagreement = len(set(directions.values())) > 1
        local_opposition = bool(item.age_specific_slopes_include_opposing_directions)
        if not disagreement and not local_opposition:
            continue
        likely_factors = []
        if directions["segmented_1990_2023"] != directions["asr_endpoint_1994_2023"]:
            likely_factors.append("calendar window and piecewise-versus-endpoint estimand")
        if directions["asr_endpoint_1994_2023"] != directions["selected_age_crude_1994_2023"]:
            likely_factors.append("age coverage, age standardization, and population weighting")
        if directions["selected_age_crude_1994_2023"] != directions["apc_global_period_slope_1994_2023"]:
            likely_factors.append("custom APC global period slope versus crude endpoint change and model constraints")
        if local_opposition:
            likely_factors.append("opposing age-specific slopes hidden by aggregate summaries")
        rows.append(
            {
                "location_name": location,
                "sex_name": sex,
                "measure_name": outcome,
                "segmented_direction_1990_2023": directions["segmented_1990_2023"],
                "asr_endpoint_direction_1994_2023": directions[
                    "asr_endpoint_1994_2023"
                ],
                "selected_age_crude_direction_1994_2023": directions[
                    "selected_age_crude_1994_2023"
                ],
                "apc_global_period_slope_label_1994_2023": directions[
                    "apc_global_period_slope_1994_2023"
                ],
                "asr_annualized_endpoint_change_1994_2023": asr_aligned,
                "selected_age_crude_annualized_change_1994_2023": crude_aligned,
                "apc_global_period_slope_1994_2023": item.apc_global_period_slope_1994_2023,
                "opposing_age_specific_slopes": local_opposition,
                "likely_explanatory_factors": "; ".join(likely_factors),
                "implementation_failure_indicated": False,
                "interpretation": (
                    "The compared methods target different age coverage, weighting, windows, "
                    "and model functions. The discrepancy is retained as a substantive "
                    "cross-method finding unless synthetic recovery or input QA fails."
                ),
            }
        )
    return pd.DataFrame(rows)


def write_methodological_notes(
    path: Path,
    contradictions: pd.DataFrame,
    age_sensitivity: pd.DataFrame,
    apc_window_sensitivity: pd.DataFrame,
) -> None:
    lines = [
        "# Internal methodological notes",
        "",
        "This file is generated from the current analysis outputs. Apparent disagreements are retained and investigated; they are not automatically described as validation failures.",
        "",
        "## Cross-method disagreements",
        "",
    ]
    if contradictions.empty:
        lines.append("No practical-label disagreements or opposing age-specific slopes were detected.")
    else:
        for row in contradictions.itertuples(index=False):
            lines.extend(
                [
                    f"### {row.location_name}, {row.sex_name}, {row.measure_name}",
                    "",
                    f"- Practical labels: segmented 1990-2023 = {row.segmented_direction_1990_2023}; ASR endpoint 1994-2023 = {row.asr_endpoint_direction_1994_2023}; selected-age crude endpoint 1994-2023 = {row.selected_age_crude_direction_1994_2023}; custom APC global period slope = {row.apc_global_period_slope_label_1994_2023}.",
                    f"- Likely factors: {row.likely_explanatory_factors}.",
                    f"- Interpretation: {row.interpretation}",
                    "",
                ]
            )
    lines.extend(["## APC window sensitivity", ""])
    changed = apc_window_sensitivity[~apc_window_sensitivity.direction_agreement]
    if changed.empty:
        lines.append("All custom APC global-period-slope practical labels agreed across the two windows.")
    else:
        for row in changed.itertuples(index=False):
            lines.append(
                f"- {row.location_name}, {row.sex_name}, {row.measure_name}: primary 1994-2023 global period slope = {row.primary_global_period_slope_1994_2023:.6f}%/year ({row.primary_direction}); sensitivity 1990-2019 slope = {row.sensitivity_global_period_slope_1990_2019:.6f}%/year ({row.sensitivity_direction}). Labels use the practical-stability band and are not equivalence tests."
            )
    lines.extend(["", "## Decomposition age-bin sensitivity", ""])
    flagged = age_sensitivity[age_sensitivity.material_age_bin_sensitivity]
    if flagged.empty:
        lines.append("No stratum met the recorded material-sensitivity rule.")
    else:
        for row in flagged.itertuples(index=False):
            lines.append(
                f"- {row.location_name}, {row.sex_name}, {row.measure_name}: maximum component shift = {row.maximum_component_shift_pct_of_total_change:.1f}% of total change; rank stable = {row.component_rank_stable}."
            )
    lines.extend(
        [
            "",
            "These accounting components are descriptive, not causal effects. Exact all-age closure is a QA property and does not prove that the source age bins are sufficiently granular for ageing attribution.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def export_nci_inputs(df: pd.DataFrame, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    trend = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR)]
    manifest = []
    for (loc, sex, outcome), p in trend.groupby(["location_name", "sex_name", "measure_name"]):
        slug = f"{loc}_{sex}_{outcome}".lower().replace(" ", "_")
        x = p.sort_values("year")[["year", "val", "lower", "upper"]].rename(columns={"val":"rate"})
        x["log_se_from_ui_sensitivity_only"] = (np.log(x.upper)-np.log(x.lower))/(2*1.96)
        x.to_csv(out_dir / f"{slug}.csv", index=False)
        manifest.append({"file":f"{slug}.csv","location":loc,"sex":sex,"outcome":outcome})
    settings = {"software":"NCI Joinpoint Regression Program","software_version":"record exact installed version at execution",
                "role":"optional validation of descriptive curve shape; not the primary inferential analysis",
                "log_transformation":True,"min_joinpoints":0,"max_joinpoints":2,
                "minimum_observations_per_segment":4,"permutations":4499,"alpha":0.05,
                "primary_error_model":"homoscedastic","sensitivity_error_model":"log SE approximated from native GBD UI",
                "inference_note":DESCRIPTIVE_INFERENCE_NOTE,
                "status":"optional; requires user registration and execution in official NCI software"}
    (out_dir/"analysis_settings.json").write_text(json.dumps(settings,indent=2),encoding="utf-8")
    pd.DataFrame(manifest).to_csv(out_dir/"input_manifest.csv",index=False)


def _consistent_software_version(values: pd.Series) -> str:
    parsed = values.astype(str).str.extract(r"(\d+\.\d+(?:\.\d+)?)", expand=False)
    if parsed.isna().any():
        raise ValueError("Every normalized NCI result row must include a parseable software version.")
    versions = parsed.unique()
    if len(versions) != 1:
        raise ValueError(f"Normalized NCI results mix software versions: {sorted(versions)}")
    return str(versions[0])


def _reference_numeric(frame: pd.DataFrame, names: tuple[str, ...]) -> pd.Series:
    for name in names:
        if name in frame:
            return pd.to_numeric(frame[name], errors="coerce")
    return pd.Series(np.nan, index=frame.index, dtype=float)


def load_normalized_nci_results(path: Path) -> dict[str, pd.DataFrame | str]:
    raw = pd.read_csv(path)
    required = {"analysis_type", "software_version", "location_name", "sex_name", "measure_name"}
    missing = required - set(raw.columns)
    if missing:
        raise ValueError(f"Normalized NCI result file is missing columns: {sorted(missing)}")
    software_version = _consistent_software_version(raw.software_version)
    trend = raw[raw.analysis_type == "trend"].copy()
    segments = raw[raw.analysis_type == "segment"].copy()
    comparisons = raw[raw.analysis_type == "comparison"].copy()
    fitted = raw[raw.analysis_type == "fitted"].copy()
    keys = ["location_name", "sex_name", "measure_name"]
    if trend[keys].drop_duplicates().shape[0] != 12:
        raise ValueError("NCI normalized results require 12 unique trend rows.")
    if segments[keys].drop_duplicates().shape[0] != 12:
        raise ValueError("NCI normalized results require segment rows for all 12 series.")
    if len(comparisons) not in (0, 12):
        raise ValueError("NCI normalized results require either zero or 12 optional pairwise comparison rows.")
    if fitted[keys + ["year"]].drop_duplicates().shape[0] != 12 * 34:
        raise ValueError("NCI normalized results require 408 fitted annual values.")
    trend_out = trend.copy()
    trend_out["software_aapc_lower_ci_reference_only"] = _reference_numeric(
        trend_out, ("aapc_lower_ci", "aapc_lower_model_ci")
    )
    trend_out["software_aapc_upper_ci_reference_only"] = _reference_numeric(
        trend_out, ("aapc_upper_ci", "aapc_upper_model_ci")
    )
    trend_out = trend_out.drop(
        columns=["aapc_lower_ci", "aapc_upper_ci", "aapc_lower_model_ci", "aapc_upper_model_ci"],
        errors="ignore",
    )
    trend_out["used_for_primary_inference"] = False
    trend_out["model_label"] = f"optional NCI Joinpoint {software_version} curve; descriptive use only"
    trend_out["inference_note"] = DESCRIPTIVE_INFERENCE_NOTE

    segment_out = segments.copy()
    segment_out["software_apc_lower_ci_reference_only"] = _reference_numeric(
        segment_out, ("apc_lower_ci", "apc_lower_model_ci")
    )
    segment_out["software_apc_upper_ci_reference_only"] = _reference_numeric(
        segment_out, ("apc_upper_ci", "apc_upper_model_ci")
    )
    segment_out = segment_out.drop(
        columns=["apc_lower_ci", "apc_upper_ci", "apc_lower_model_ci", "apc_upper_model_ci"],
        errors="ignore",
    )
    segment_out["used_for_primary_inference"] = False
    segment_out["inference_note"] = DESCRIPTIVE_INFERENCE_NOTE

    pair_out = comparisons.copy()
    pair_out["software_parallelism_p_value_reference_only"] = _reference_numeric(
        pair_out, ("parallelism_p_value", "p_value")
    )
    pair_out = pair_out.drop(columns=["parallelism_p_value", "p_value"], errors="ignore")
    pair_out["used_for_primary_inference"] = False
    pair_out["method_note"] = DESCRIPTIVE_INFERENCE_NOTE
    return {
        "summary": trend_out,
        "segments": segment_out,
        "pairwise": pair_out,
        "fitted": fitted,
        "software_version": software_version,
    }


def plot_asr(df: pd.DataFrame, path: Path) -> None:
    trend = df[(df.measure_name.isin(OUTCOMES)) & (df.metric_name == "Rate") & (df.age_name == ASR)]
    fig, axes = plt.subplots(1,3,figsize=(14,4.5),sharex=True)
    for ax,outcome in zip(axes,OUTCOMES):
        for loc in LOCATIONS:
            for sex in SEXES:
                p=trend[(trend.location_name==loc)&(trend.sex_name==sex)&(trend.measure_name==outcome)].sort_values("year")
                ax.plot(p.year,p.val,color=COLORS[loc],ls=SEX_LINE[sex],lw=2,label=f"{loc.replace('United States of America','United States')} {sex}")
                ax.fill_between(p.year,p.lower,p.upper,color=COLORS[loc],alpha=.08)
        ax.set_title(outcome); ax.set_xlabel("Year"); ax.set_ylabel("ASR per 100,000"); ax.grid(alpha=.2)
    handles,labels=axes[0].get_legend_handles_labels(); fig.legend(handles,labels,loc="lower center",ncol=4,frameon=False)
    fig.suptitle("Sex-specific schizophrenia age-standardized rates, 1990-2023",fontweight="bold")
    fig.tight_layout(rect=(0,.12,1,.93)); fig.savefig(path,dpi=300,bbox_inches="tight"); plt.close(fig)


def plot_segmented(df: pd.DataFrame, fitted: pd.DataFrame, path: Path) -> None:
    fig,axes=plt.subplots(3,2,figsize=(12,11),sharex=True)
    for r,outcome in enumerate(OUTCOMES):
        for c,sex in enumerate(SEXES):
            ax=axes[r,c]
            for loc in LOCATIONS:
                p=fitted[(fitted.location_name==loc)&(fitted.sex_name==sex)&(fitted.measure_name==outcome)].sort_values("year")
                ax.scatter(p.year,p.val,s=8,color=COLORS[loc],alpha=.45)
                ax.plot(p.year,p.fitted,color=COLORS[loc],lw=2,label=loc.replace("United States of America","United States"))
            ax.set_title(f"{outcome}: {sex}"); ax.set_ylabel("ASR per 100,000"); ax.grid(alpha=.2)
    for ax in axes[-1]: ax.set_xlabel("Year")
    h,l=axes[0,0].get_legend_handles_labels(); fig.legend(h,l,loc="lower center",ncol=2,frameon=False)
    fig.suptitle("Descriptive BIC-selected segmented rate trends",fontweight="bold")
    fig.tight_layout(rect=(0,.04,1,.96)); fig.savefig(path,dpi=300,bbox_inches="tight"); plt.close(fig)


def plot_age_patterns(df: pd.DataFrame, path: Path) -> None:
    ages = select_decomposition_ages(set(df.age_name.dropna().astype(str)))
    p=df[(df.measure_name.isin(OUTCOMES))&(df.metric_name=="Rate")&df.age_name.isin(ages)&(df.year==2023)].copy()
    p["age_index"]=p.age_name.map({x:i for i,x in enumerate(ages)})
    fig,axes=plt.subplots(3,2,figsize=(12,11),sharex=True)
    for r,outcome in enumerate(OUTCOMES):
        for c,loc in enumerate(LOCATIONS):
            ax=axes[r,c]
            for sex in SEXES:
                s=p[(p.measure_name==outcome)&(p.location_name==loc)&(p.sex_name==sex)].sort_values("age_index")
                ax.plot(s.age_index,s.val,marker="o",ls=SEX_LINE[sex],lw=2,label=sex)
                ax.fill_between(s.age_index,s.lower,s.upper,alpha=.12)
            ax.set_title(f"{outcome}: {loc.replace('United States of America','United States')}"); ax.set_ylabel("Rate per 100,000"); ax.grid(alpha=.2)
    for ax in axes[-1]: ax.set_xticks(range(len(ages))); ax.set_xticklabels([x.replace(" years","") for x in ages],rotation=45,ha="right")
    h,l=axes[0,0].get_legend_handles_labels(); fig.legend(h,l,loc="lower center",ncol=2,frameon=False)
    fig.suptitle("Age-specific schizophrenia burden in 2023",fontweight="bold"); fig.tight_layout(rect=(0,.06,1,.96)); fig.savefig(path,dpi=300,bbox_inches="tight"); plt.close(fig)


def plot_decomposition(decomp: pd.DataFrame, path: Path) -> None:
    d=decomp[(decomp.start_year==1990)&(decomp.end_year==2023)]
    fig,axes=plt.subplots(3,2,figsize=(12,10))
    x = np.arange(2)
    width = .22
    for r,outcome in enumerate(OUTCOMES):
        for c,sex in enumerate(SEXES):
            ax = axes[r,c]
            p = d[(d.measure_name==outcome)&(d.sex_name==sex)].set_index("location_name").reindex(LOCATIONS)
            for j,(component,color) in enumerate(COMPONENT_COLORS.items()):
                ax.bar(x+(j-1)*width,p[component],width,label=component.replace("_"," ").title(),color=color)
            ax.axhline(0,color="black",lw=.7)
            ax.set_xticks(x)
            ax.set_xticklabels(["China","United States"])
            ax.set_title(f"{outcome}: {sex}")
            ax.set_ylabel("Change in all-age quantity")
            ax.grid(axis="y",alpha=.2)
    h,l = axes[0,0].get_legend_handles_labels()
    fig.legend(h,l,loc="lower center",ncol=3,frameon=False)
    fig.suptitle("Shapley decomposition of all-age burden change, 1990-2023",fontweight="bold")
    fig.tight_layout(rect=(0,.05,1,.96))
    fig.savefig(path,dpi=300,bbox_inches="tight")
    plt.close(fig)


def plot_counts(df: pd.DataFrame, path: Path) -> None:
    p=df[(df.measure_name.isin(OUTCOMES))&(df.metric_name=="Number")&(df.age_name==ALL_AGES)]
    fig,axes=plt.subplots(1,3,figsize=(14,4.5),sharex=True)
    for ax,outcome in zip(axes,OUTCOMES):
        for loc in LOCATIONS:
            for sex in SEXES:
                s=p[(p.measure_name==outcome)&(p.location_name==loc)&(p.sex_name==sex)].sort_values("year")
                ax.plot(s.year,s.val,color=COLORS[loc],ls=SEX_LINE[sex],lw=2,label=f"{loc.replace('United States of America','United States')} {sex}")
        ax.set_title(outcome); ax.set_ylabel("All-age count"); ax.set_xlabel("Year"); ax.grid(alpha=.2)
    h,l=axes[0].get_legend_handles_labels(); fig.legend(h,l,loc="lower center",ncol=4,frameon=False); fig.tight_layout(rect=(0,.12,1,1)); fig.savefig(path,dpi=300,bbox_inches="tight"); plt.close(fig)


def plot_apc(apc: dict[str,pd.DataFrame], path: Path) -> None:
    specifications = [
        ("local_drift", "age_midpoint", "local_drift", "Age-specific log-rate slope"),
        ("age_curve", "age_midpoint", "longitudinal_age_rr", "Descriptive age rate ratio"),
        ("period_rr", "period_midpoint", "period_rr", "Period-curvature rate ratio"),
        ("cohort_rr", "cohort_midpoint", "cohort_rr", "Cohort-curvature rate ratio"),
    ]
    fig, axes = plt.subplots(len(OUTCOMES), len(specifications), figsize=(16, 11))
    for row_index, outcome in enumerate(OUTCOMES):
        for column_index, (key, xcol, ycol, title) in enumerate(specifications):
            ax = axes[row_index, column_index]
            data = apc[key]
            for loc in LOCATIONS:
                for sex in SEXES:
                    s = data[
                        data.location_name.eq(loc)
                        & data.sex_name.eq(sex)
                        & data.measure_name.eq(outcome)
                    ].sort_values(xcol)
                    ax.plot(
                        s[xcol], s[ycol], color=COLORS[loc], ls=SEX_LINE[sex], lw=1.8,
                        label=f"{loc.replace('United States of America','United States')} {sex}",
                    )
            ax.axhline(0 if key == "local_drift" else 1, color="black", lw=.6)
            if row_index == 0:
                ax.set_title(title)
            if column_index == 0:
                ax.set_ylabel(f"{outcome}\n%/year")
            else:
                ax.set_ylabel(f"{outcome}\nrate ratio")
            ax.grid(alpha=.2)
    h, l = axes[0, 0].get_legend_handles_labels()
    fig.legend(h, l, loc="lower center", ncol=4, frameon=False)
    fig.suptitle("Custom descriptive age-period-cohort summaries", fontweight="bold")
    fig.tight_layout(rect=(0, .05, 1, .96))
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def write_tables(tables: dict[str, pd.DataFrame], out: Path) -> None:
    out.mkdir(parents=True, exist_ok=True)
    # Remove only fixed filenames generated by earlier schemas. Unknown CSVs are preserved.
    for filename in OBSOLETE_TABLE_CSVS:
        (out / filename).unlink(missing_ok=True)
    # Optional NCI tables must not survive a later build that did not import NCI output.
    for name in OPTIONAL_NCI_TABLES - tables.keys():
        (out / f"{name}.csv").unlink(missing_ok=True)
    for name, table in tables.items():
        table.to_csv(out / f"{name}.csv", index=False)
    # The preferred artifact-tool runtime is not available in every noninteractive
    # build environment.  This deterministic fallback keeps the workbook useful,
    # auditable, and visually consistent while CSV remains the canonical output.
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    def safe_sheet_name(name: str, used: set[str]) -> str:
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "Sheet"
        candidate = cleaned
        suffix = 1
        while candidate.casefold() in used:
            suffix += 1
            tail = f"_{suffix}"
            candidate = f"{cleaned[:31-len(tail)]}{tail}"
        used.add(candidate.casefold())
        return candidate

    workbook_path = out / "analysis_tables.xlsx"
    used_names = {"readme", "key results"}
    sheet_map = {name: safe_sheet_name(name, used_names) for name in tables}
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                ["Analysis tables"],
                ["Canonical data", "CSV files beside this workbook"],
                ["Workbook purpose", "Review and navigation; no inferential formulas"],
                ["Uncertainty", "Native GBD UIs only; derived analyses are point estimates"],
                ["Validation", "See build_metadata.json and qa/validation_summary.json"],
                [],
                ["Table", "Worksheet", "Rows", "Columns"],
                *[
                    [name, sheet_map[name], len(table), len(table.columns)]
                    for name, table in tables.items()
                ],
            ]
        ).to_excel(writer, sheet_name="README", index=False, header=False)
        if "cross_analysis_consistency" in tables:
            key_results = tables["cross_analysis_consistency"][
                [
                    "location_name",
                    "sex_name",
                    "measure_name",
                    "asr_endpoint_percent_change_1990_2023",
                    "all_age_count_percent_change_1990_2023",
                    "segmented_aapc_1990_2023",
                    "segmented_overall_direction",
                    "population_size_change",
                    "age_structure_change",
                    "age_specific_rate_change",
                    "apc_global_period_slope_1994_2023",
                ]
            ].rename(
                columns={
                    "location_name": "Location",
                    "sex_name": "Sex",
                    "measure_name": "Outcome",
                    "asr_endpoint_percent_change_1990_2023": "ASR change, %",
                    "all_age_count_percent_change_1990_2023": "Count change, %",
                    "segmented_aapc_1990_2023": "Segmented AAPC, %",
                    "segmented_overall_direction": "Direction",
                    "population_size_change": "Population size",
                    "age_structure_change": "Age structure",
                    "age_specific_rate_change": "Age-specific rate",
                    "apc_global_period_slope_1994_2023": "APC global period slope, %",
                }
            )
            key_results.to_excel(writer, sheet_name="Key Results", index=False)
        for name, table in tables.items():
            table.to_excel(writer, sheet_name=sheet_map[name], index=False)

        workbook = writer.book
        navy = "1F4E78"; teal = "0F6B78"; light = "D9EAF7"; pale = "F2F4F7"
        warning = "FCE4D6"; good = "E2F0D9"
        readme = workbook["README"]
        readme.sheet_view.showGridLines = False
        readme.freeze_panes = "A7"
        readme.merge_cells("A1:D1")
        readme["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
        readme["A1"].fill = PatternFill("solid", fgColor=navy)
        readme["A1"].alignment = Alignment(vertical="center")
        readme.row_dimensions[1].height = 30
        for row in range(2, 6):
            readme.cell(row, 1).font = Font(name="Calibri", bold=True, color=navy)
            readme.cell(row, 2).alignment = Alignment(wrap_text=True)
        for cell in readme[7]:
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=teal)
        readme.column_dimensions["A"].width = 39
        readme.column_dimensions["B"].width = 36
        readme.column_dimensions["C"].width = 12
        readme.column_dimensions["D"].width = 12
        readme.page_setup.orientation = "landscape"
        readme.sheet_properties.pageSetUpPr.fitToPage = True
        readme.page_setup.fitToWidth = 1
        readme.page_setup.fitToHeight = 1
        readme.print_area = f"A1:D{readme.max_row}"

        data_sheets = (["Key Results"] if "Key Results" in workbook.sheetnames else []) + list(sheet_map.values())
        for index, sheet_name in enumerate(data_sheets, start=1):
            sheet = workbook[sheet_name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            # Keep printed tables legible: wide canonical tables flow across
            # multiple horizontal pages instead of being reduced to tiny type.
            sheet.page_setup.fitToWidth = max(1, math.ceil(sheet.max_column / 12))
            sheet.page_setup.fitToHeight = 1
            sheet.page_setup.orientation = "landscape"
            sheet.print_title_rows = "1:1"
            sheet.sheet_properties.tabColor = (
                "70AD47" if "audit" in sheet_name or "validation" in sheet_name
                else "ED7D31" if "sensitivity" in sheet_name
                else "7F8C8D" if "provenance" in sheet_name or "source" in sheet_name
                else teal
            )
            for cell in sheet[1]:
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            sheet.row_dimensions[1].height = 32
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Calibri", size=9)
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
            for column_index, column_cells in enumerate(sheet.iter_cols(), start=1):
                observed = max(
                    (len(str(cell.value)) for cell in list(column_cells)[:250] if cell.value is not None),
                    default=8,
                )
                sheet.column_dimensions[get_column_letter(column_index)].width = min(max(observed + 2, 9), 34)
            if sheet.max_row >= 2 and sheet.max_column >= 1:
                table_name = f"T{index}_{re.sub(r'[^A-Za-z0-9_]', '_', sheet_name)[:200]}"
                excel_table = Table(displayName=table_name, ref=sheet.dimensions)
                excel_table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False,
                )
                sheet.add_table(excel_table)
                for column_index, header_cell in enumerate(sheet[1], start=1):
                    header = str(header_cell.value).casefold()
                    positive_boolean = any(
                        token in header
                        for token in (
                            "stable",
                            "complete",
                            "ready",
                            "within_tolerance",
                            "numerically_identical",
                        )
                    )
                    negative_boolean = "failure_indicated" in header
                    if positive_boolean or negative_boolean:
                        col = get_column_letter(column_index)
                        false_fill = good if negative_boolean else warning
                        true_fill = warning if negative_boolean else good
                        sheet.conditional_formatting.add(
                            f"{col}2:{col}{sheet.max_row}",
                            CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor=false_fill)),
                        )
                        sheet.conditional_formatting.add(
                            f"{col}2:{col}{sheet.max_row}",
                            CellIsRule(operator="equal", formula=["TRUE"], fill=PatternFill("solid", fgColor=true_fill)),
                        )
            sheet.print_area = (
                f"A1:{get_column_letter(sheet.max_column)}{min(sheet.max_row, 30)}"
            )


def write_data_dictionary(tables: dict[str, pd.DataFrame], path: Path) -> None:
    descriptions = {
        "val":"GBD posterior mean point estimate", "lower":"Lower native 95% GBD uncertainty bound",
        "upper":"Upper native 95% GBD uncertainty bound", "aapc":"Average annual percentage change",
        "population_source":"Provenance status of population denominator", "closure_error":"Component sum minus total reconstructed change",
        "all_age_count_start_reconstructed":"Reconstructed all-age count at the start year",
        "all_age_count_end_reconstructed":"Reconstructed all-age count at the end year",
        "reconstructed_all_age_count":"Sum of reconstructed counts across the complete age partition",
        "reported_all_age_count":"Reported GBD all-age count",
        "formal_inference_performed":"Whether hypothesis tests or confidence intervals were treated as valid for GBD estimates",
        "lag1_residual_autocorrelation":"Descriptive lag-one correlation of fitted log-rate residuals",
        "durbin_watson":"Descriptive Durbin-Watson residual statistic; not used for inference",
    }
    rows=[]
    for table_name,frame in tables.items():
        for col in frame.columns:
            if col in {"lower", "upper"}:
                uncertainty_class = "native GBD UI"
            elif "reference_only" in col:
                uncertainty_class = "conditional software result; reference only"
            else:
                uncertainty_class = "none/not applicable"
            rows.append({"table":table_name,"column":col,"dtype":str(frame[col].dtype),
                         "description":descriptions.get(col,col.replace("_"," ").capitalize()),
                         "uncertainty_class":uncertainty_class})
    pd.DataFrame(rows).to_csv(path,index=False)


def portable_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return resolved.relative_to(ROOT).as_posix()
    except ValueError:
        return f"external/{resolved.name}"


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_export_metadata(path: Path, expected_role: str) -> dict:
    """Validate a provenance sidecar and the hashes of its preserved raw files."""
    try:
        metadata = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Cannot read export metadata {path}: {exc}") from exc
    required = {
        "export_role",
        "gbd_release",
        "retrieval_date",
        "export_id",
        "source_url",
        "query_dimensions",
        "raw_files",
    }
    missing = required - set(metadata)
    if missing:
        raise ValueError(f"Export metadata is missing fields: {sorted(missing)}")
    if metadata["export_role"] != expected_role:
        raise ValueError(
            f"Export metadata role must be {expected_role!r}; found {metadata['export_role']!r}."
        )
    if "2023" not in str(metadata["gbd_release"]):
        raise ValueError("Export metadata must identify GBD 2023.")
    try:
        date.fromisoformat(str(metadata["retrieval_date"]))
    except ValueError as exc:
        raise ValueError("Export metadata retrieval_date must use YYYY-MM-DD.") from exc
    export_id = str(metadata["export_id"]).strip()
    if not export_id or "identifier" in export_id.casefold() or "replace" in export_id.casefold():
        raise ValueError("Export metadata export_id must be non-empty.")
    source_url = str(metadata["source_url"])
    if not re.match(r"https://([a-z0-9-]+\.)*healthdata\.org(?:/|$)", source_url, re.I):
        raise ValueError("Export metadata source_url must be an official HTTPS healthdata.org URL.")
    if not isinstance(metadata["query_dimensions"], dict) or not metadata["query_dimensions"]:
        raise ValueError("Export metadata query_dimensions must be a non-empty object.")
    dimensions = metadata["query_dimensions"]
    required_dimensions = {"locations", "sexes", "years", "ages", "measures", "metrics"}
    missing_dimensions = required_dimensions - set(dimensions)
    if missing_dimensions:
        raise ValueError(
            f"Export metadata query_dimensions is missing: {sorted(missing_dimensions)}"
        )

    def dimension_set(name: str) -> set[str]:
        values = dimensions[name]
        if not isinstance(values, list) or not values:
            raise ValueError(f"Export metadata query_dimensions.{name} must be a non-empty list.")
        return {str(value).strip() for value in values}

    if dimension_set("locations") != set(LOCATIONS):
        raise ValueError("Export metadata locations must exactly match the two analysis locations.")
    if dimension_set("sexes") != set(SEXES):
        raise ValueError("Export metadata sexes must exactly match Female and Male.")
    years = dimensions["years"]
    valid_years = (
        str(years).replace("–", "-") == "1990-2023"
        or (
            isinstance(years, list)
            and {int(value) for value in years} == set(YEARS)
        )
    )
    if not valid_years:
        raise ValueError("Export metadata years must cover every year from 1990 through 2023.")

    ages = dimension_set("ages")
    measures = dimension_set("measures")
    metrics = dimension_set("metrics")
    if expected_role == "burden":
        required_ages = {*FINE_DECOMPOSITION_AGES, ALL_AGES, ASR}
        if not required_ages <= ages:
            raise ValueError("Burden export metadata is missing required fine or summary ages.")
        if not set(OUTCOMES) <= measures:
            raise ValueError("Burden export metadata is missing a required study outcome.")
        if not {"Number", "Rate"} <= metrics:
            raise ValueError("Burden export metadata must include Number and Rate metrics.")
    elif expected_role == "population":
        if ages != set(FINE_DECOMPOSITION_AGES):
            raise ValueError("Population export metadata ages must exactly match the 20 fine groups.")
        if "Population" not in measures or "Number" not in metrics:
            raise ValueError("Population export metadata must identify Population and Number.")
    if not isinstance(metadata["raw_files"], list) or not metadata["raw_files"]:
        raise ValueError("Export metadata raw_files must be a non-empty list.")
    for item in metadata["raw_files"]:
        if not isinstance(item, dict) or {"file", "sha256"} - set(item):
            raise ValueError("Each raw_files entry must contain file and sha256.")
        raw_relative = Path(str(item["file"]))
        if raw_relative.is_absolute():
            raise ValueError("Raw export paths in metadata must be repository-relative.")
        raw_path = ROOT / raw_relative
        if not raw_path.is_file():
            raise ValueError(f"Preserved raw export is missing: {raw_relative.as_posix()}")
        expected_hash = str(item["sha256"]).lower()
        if not re.fullmatch(r"[0-9a-f]{64}", expected_hash):
            raise ValueError(f"Invalid SHA-256 in metadata for {raw_relative.as_posix()}.")
        actual_hash = file_sha256(raw_path)
        if actual_hash != expected_hash:
            raise ValueError(
                f"Raw export hash mismatch for {raw_relative.as_posix()}: "
                f"expected {expected_hash}, found {actual_hash}."
            )
    metadata = dict(metadata)
    metadata["metadata_file"] = portable_path(path)
    metadata["metadata_sha256"] = file_sha256(path)
    return metadata


def provenance_table(
    burden_path: Path,
    population_path: Path | None,
    population_source: str,
    burden_hash: str,
    population_hash: str | None,
    burden_metadata: dict | None = None,
    population_metadata: dict | None = None,
) -> pd.DataFrame:
    def dimensions(metadata: dict | None, fallback: str) -> str:
        if not metadata:
            return fallback
        return "; ".join(
            f"{key}={json.dumps(value, ensure_ascii=False, sort_keys=True)}"
            for key, value in sorted(metadata["query_dimensions"].items())
        )

    population_file = (
        portable_path(population_path)
        if population_path
        else "derived from burden count/rate pairs"
    )
    return pd.DataFrame([
        {
            "source_role": "burden estimates",
            "gbd_release": "GBD 2023",
            "file": portable_path(burden_path),
            "sha256": burden_hash,
            "retrieval_date": (
                burden_metadata["retrieval_date"]
                if burden_metadata
                else "not recorded in source archive"
            ),
            "export_id": burden_metadata["export_id"] if burden_metadata else "not recorded",
            "metadata_file": burden_metadata["metadata_file"] if burden_metadata else "not supplied",
            "metadata_sha256": burden_metadata["metadata_sha256"] if burden_metadata else "not applicable",
            "dimensions": dimensions(
                burden_metadata,
                "China; United States; 1990-2023; Female; Male; incidence; "
                "prevalence; YLD; DALY; counts and rates",
            ),
            "status": "included",
        },
        {
            "source_role": "population denominators",
            "gbd_release": (
                "GBD 2023"
                if population_source == "official_GBD_2023"
                else "UNVERIFIED - PROVISIONAL ONLY"
            ),
            "file": population_file,
            "sha256": population_hash if population_hash else "not applicable - derived",
            "retrieval_date": (
                population_metadata["retrieval_date"]
                if population_metadata
                else "not supplied"
            ),
            "export_id": population_metadata["export_id"] if population_metadata else "not supplied",
            "metadata_file": population_metadata["metadata_file"] if population_metadata else "not supplied",
            "metadata_sha256": population_metadata["metadata_sha256"] if population_metadata else "not applicable",
            "dimensions": dimensions(
                population_metadata,
                "China; United States; 1990-2023; Female; Male; "
                + (
                    "five-year ages 0-4 through 90-94 plus 95+"
                    if population_source == "official_GBD_2023"
                    else "provisional ages 0-14 through 70+"
                ),
            ),
            "status": population_source,
        },
        {
            "source_role": "GBD percent metric",
            "gbd_release": "GBD 2023",
            "file": portable_path(burden_path),
            "sha256": burden_hash,
            "retrieval_date": "not applicable",
            "export_id": "not applicable",
            "metadata_file": "not applicable",
            "metadata_sha256": "not applicable",
            "dimensions": "mixed measure-specific denominators",
            "status": "excluded",
        },
        {
            "source_role": "probability of death",
            "gbd_release": "GBD 2023",
            "file": "prepared_inputs/GBD_1990_2023_ProbabilityOfDeath_ChinaUS_Schizophrenia.csv",
            "sha256": "not computed - excluded source",
            "retrieval_date": "not applicable",
            "export_id": "not applicable",
            "metadata_file": "not applicable",
            "metadata_sha256": "not applicable",
            "dimensions": "not a schizophrenia-specific causal outcome",
            "status": "excluded",
        },
        {
            "source_role": "risk factor extract",
            "gbd_release": "GBD 2023",
            "file": "schizo/IHME-GBD_2023_DATA-5ef7a575-1.zip",
            "sha256": "not computed - excluded source",
            "retrieval_date": "not applicable",
            "export_id": "not applicable",
            "metadata_file": "not applicable",
            "metadata_sha256": "not applicable",
            "dimensions": "single sexual-violence risk branch",
            "status": "excluded",
        },
    ])


def run(args) -> dict:
    out=Path(args.output_dir)
    if out.exists() and any(out.iterdir()):
        raise ValueError(
            f"Analysis output directory must be absent or empty to prevent stale artifacts: {out}"
        )
    tables_dir=out/"tables"; main_fig=out/"figures"/"main"; supp_fig=out/"figures"/"supplement"
    for p in (tables_dir,main_fig,supp_fig,out/"qa",out/"nci_joinpoint_inputs"): p.mkdir(parents=True,exist_ok=True)
    burden_path=Path(args.burden_csv); population_path=Path(args.population_csv) if args.population_csv else None
    burden_metadata_path=Path(args.burden_metadata_json) if getattr(args,"burden_metadata_json",None) else None
    population_metadata_path=Path(args.population_metadata_json) if getattr(args,"population_metadata_json",None) else None
    burden_hash=file_sha256(burden_path); population_hash=file_sha256(population_path) if population_path else None
    burden_metadata=load_export_metadata(burden_metadata_path,"burden") if burden_metadata_path else None
    population_metadata=load_export_metadata(population_metadata_path,"population") if population_metadata_path else None
    burden=load_burden(burden_path)
    validate_required_burden_summaries(burden)
    age_granularity=burden_age_granularity_audit(burden)
    if population_path:
        if burden_metadata is None or population_metadata is None:
            raise ValueError(
                "A production build requires --burden-metadata-json and "
                "--population-metadata-json with verified raw-file hashes."
            )
        pop=load_official_population(population_path,args.population_release)
        if not bool(age_granularity.fine_age_panel_complete.all()):
            raise ValueError(
                "Official production population was supplied, but the burden input lacks "
                "complete fine-age Number and Rate panels for all outcomes. Obtain the "
                "matching fine-age GBD 2023 burden export before a production build."
            )
    elif args.allow_proxy_population:
        pop=infer_proxy_population(burden)
    else:
        raise SystemExit("A matching official GBD 2023 --population-csv is required. Use --allow-proxy-population only for a visibly provisional build.")
    validate_population(pop)
    reconstructed_population=infer_proxy_population(
        burden,
        allow_undefined=pop.population_source.iloc[0] == "official_GBD_2023",
    )
    if pop.population_source.iloc[0]=="official_GBD_2023":
        population_comparison=compare_population_sources(pop,reconstructed_population)
    else:
        population_comparison=pd.DataFrame(columns=[
            "location_name","sex_name","age_name","year","official_population",
            "reconstructed_population","absolute_difference","relative_difference_pct",
            "key_match_status","reconstruction_available","comparison_status",
        ])
    audit,duplicate,reconstruction=audit_burden(burden,pop)
    source_zero_audit,source_zero_validation=audit_source_export_zeros(
        burden,burden_metadata_path
    )
    all_age_reconstruction=all_age_count_reconstruction(burden,reconstruction)
    yld_identity=verify_yld_daly_identity(burden)
    endpoints=endpoint_table(burden); country,sex=contrast_tables(burden)
    seg,segments,fitted=run_segmented(burden)
    ar1_sensitivity,ar1_segments=segmented_ar1_sensitivity(burden,seg)
    stability_sensitivity=practical_stability_sensitivity(seg)
    segmented_sensitivity=segmented_specification_sensitivity(burden)
    weighted=weighted_trend_sensitivity(burden,seg)
    excluding_2020_2023,_,_=run_segmented(burden,end_year=2019)
    pairwise=build_trajectory_contrasts(burden)
    nci_valid=False; nci_version=None; nci_tables={}
    if args.nci_results_csv:
        official=load_normalized_nci_results(Path(args.nci_results_csv))
        official_fitted=official["fitted"].merge(
            burden[(burden.metric_name=="Rate")&(burden.age_name==ASR)&burden.measure_name.isin(OUTCOMES)][["location_name","sex_name","measure_name","year","val","lower","upper"]],
            on=["location_name","sex_name","measure_name","year"],how="left",validate="one_to_one")
        nci_tables={"nci_validation_summary":official["summary"],"nci_validation_segments":official["segments"],
                    "nci_validation_comparisons":official["pairwise"],"nci_validation_fitted":official_fitted}
        nci_version=str(official["software_version"]); nci_valid=True
    decomp=run_decomposition(burden,pop); annual=chained_decomposition(burden,pop,1); fiveyear=chained_decomposition(burden,pop,5)
    decomp_path_sensitivity=decomposition_path_sensitivity_summary(
        decomp,annual,fiveyear
    )
    supported_age_incidence_decomp=incidence_supported_age_decomposition(burden,pop)
    decomp_age_sensitivity=decomposition_age_bin_sensitivity(burden,pop)
    apc=run_secondary_apc(burden,pop,True)
    apc_unweighted=run_secondary_apc(burden,pop,True,weighting="equal")
    apc_weighting_sensitivity=compare_apc_weighting(apc,apc_unweighted)
    apc_pre=run_secondary_apc(burden,pop,False)
    apc_agreement=compare_primary_apc_directions(burden,seg,apc["summary"])
    apc_window_sensitivity=compare_apc_windows(apc["summary"],apc_pre["summary"])
    cross_analysis=build_cross_analysis_consistency(
        burden,pop,endpoints,seg,apc,decomp
    )
    contradictions=investigate_cross_method_contradictions(
        burden,pop,cross_analysis
    )
    provenance=provenance_table(
        burden_path,population_path,pop.population_source.iloc[0],burden_hash,
        population_hash,burden_metadata,population_metadata
    )
    tables={"data_audit":audit,"duplicate_audit":duplicate,"burden_age_granularity_audit":age_granularity,
            "source_export_zero_audit":source_zero_audit,
            "population_reconstruction":reconstruction,"population_source_comparison":population_comparison,
            "all_age_count_reconstruction":all_age_reconstruction,"yld_daly_identity":yld_identity,
            "endpoint_summary":endpoints,"country_contrasts":country,"sex_contrasts":sex,"segmented_summary":seg,"segmented_segments":segments,
            "segmented_ar1_sensitivity":ar1_sensitivity,
            "segmented_ar1_segments":ar1_segments,
            "trend_practical_stability_sensitivity":stability_sensitivity,
            "segmented_fitted":fitted,"segmented_specification_sensitivity":segmented_sensitivity,
            "trajectory_contrasts":pairwise,"ui_weighted_sensitivity":weighted,
            "trend_excluding_2020_2023":excluding_2020_2023,
            "decomposition":decomp,"decomposition_age_bin_sensitivity":decomp_age_sensitivity,
            "incidence_supported_age_decomposition":supported_age_incidence_decomp,
            "annual_chained_decomposition":annual,"fiveyear_chained_decomposition":fiveyear,
            "decomposition_path_sensitivity":decomp_path_sensitivity,
            **format_apc_tables(apc,"apc_descriptive"),
            **format_apc_tables(apc_unweighted,"apc_unweighted"),
            "apc_weighting_sensitivity":apc_weighting_sensitivity,
            **format_apc_tables(apc_pre,"apc_sensitivity_1990_2019"),
            "apc_window_sensitivity":apc_window_sensitivity,
            "apc_primary_direction_agreement":apc_agreement,
            "cross_analysis_consistency":cross_analysis,
            "cross_method_contradictions":contradictions,
            "provenance":provenance,**nci_tables}
    write_tables(tables,tables_dir); write_data_dictionary(tables,out/"data_dictionary.csv"); export_nci_inputs(burden,out/"nci_joinpoint_inputs")
    write_methodological_notes(
        out/"qa"/"methodological_notes.md",contradictions,decomp_age_sensitivity,
        apc_window_sensitivity
    )
    plot_asr(burden,main_fig/"figure_1_asr_trends.png"); plot_segmented(burden,fitted,main_fig/"figure_2_segmented_trends.png")
    plot_age_patterns(burden,main_fig/"figure_3_age_patterns.png"); plot_decomposition(decomp,main_fig/"figure_4_decomposition.png")
    plot_counts(burden,supp_fig/"figure_s1_counts.png"); plot_apc(apc,supp_fig/"figure_s2_custom_apc_summaries.png")
    all_age_reconstruction_valid=bool(len(all_age_reconstruction)==len(LOCATIONS)*len(SEXES)*len(OUTCOMES)*len(YEARS)
                                      and all_age_reconstruction.within_tolerance.all())
    internal_validation_passed=bool((audit.all_age_count_years.eq(34)&audit.asr_years.eq(34)).all()
                                    and int(duplicate.duplicate_dimensional_keys.iloc[0])==0
                                    and int(audit.invalid_ui_rows.sum())==0 and int(audit.negative_rows.sum())==0
                                    and source_zero_validation["expected_age_outcome_pattern"]
                                    and source_zero_validation["complete_location_sex_year_metric_pattern"]
                                    and (source_zero_validation["source_export_provenance_verified"] or burden_metadata_path is None)
                                    and bool(yld_identity.audit_passed.iloc[0])
                                    and all_age_reconstruction_valid and float(decomp.closure_error.abs().max())<1e-8)
    trend_status="descriptive BIC-selected segmented curves; no formal trend inference"
    if nci_valid: trend_status+=f"; optional NCI Joinpoint {nci_version} validation imported"
    fine_age_valid=bool(age_granularity.fine_age_panel_complete.all())
    source_metadata_complete=bool(burden_metadata and population_metadata)
    decomposition_ages=select_decomposition_ages(set(pop.age_name.dropna().astype(str)))
    data_readiness_passed=bool(pop.population_source.iloc[0]=="official_GBD_2023" and fine_age_valid and source_metadata_complete and internal_validation_passed)
    metadata={"build_date":date.today().isoformat(),"population_status":pop.population_source.iloc[0],
              "fine_age_burden_validated":fine_age_valid,
              "decomposition_age_partition":"fine_5_year" if decomposition_ages==FINE_DECOMPOSITION_AGES else "provisional_broad_end_groups",
              "source_metadata_complete":source_metadata_complete,
              "data_readiness_passed":data_readiness_passed,
              "analysis_ready":data_readiness_passed,
              "trend_status":trend_status,"trend_selection_method":"deterministic BIC; zero to two breakpoints",
              "formal_trend_inference_performed":False,"official_nci_results_optional":True,
              "official_nci_software_version":nci_version,
              "burden_csv":portable_path(burden_path),"population_csv":portable_path(population_path) if population_path else None,
              "burden_metadata_json":portable_path(burden_metadata_path) if burden_metadata_path else None,
              "population_metadata_json":portable_path(population_metadata_path) if population_metadata_path else None,
              "burden_csv_sha256":burden_hash,"population_csv_sha256":population_hash,
              "limitations":[
                  "No posterior draws","Cross-year GBD correlation unavailable",
                  *(["Provisional burden input combines ages 0-14 and 70+ into broad groups"] if not fine_age_valid else []),
                  "Ecological modeled estimates","No causal health-system inference"
              ]}
    (out/"build_metadata.json").write_text(json.dumps(metadata,indent=2),encoding="utf-8")
    validation={
        "all_primary_panels_complete_34_years":bool((audit.all_age_count_years.eq(34)&audit.asr_years.eq(34)).all()),
        "duplicate_dimensional_keys":int(duplicate.duplicate_dimensional_keys.iloc[0]),
        "invalid_ui_rows":int(audit.invalid_ui_rows.sum()),
        "negative_rows":int(audit.negative_rows.sum()),
        "zero_rows":int(audit.zero_rows.sum()),
        "nonpositive_rows":int(audit.nonpositive_rows.sum()),
        "source_export_zero_audit":source_zero_validation,
        "yld_daly_audit_status":str(yld_identity.audit_status.iloc[0]),
        "yld_daly_audit_passed":bool(yld_identity.audit_passed.iloc[0]),
        "yld_daly_numerically_identical":bool(yld_identity.numerically_identical.iloc[0]),
        "population_reconstruction_p99_absolute_relative_error_pct":float(reconstruction.relative_error_pct.abs().quantile(.99)),
        "all_age_count_reconstruction_rows":int(len(all_age_reconstruction)),
        "all_age_count_reconstruction_max_absolute_relative_error_pct":float(all_age_reconstruction.relative_error_pct.abs().max()),
        "all_age_count_reconstruction_within_tolerance":all_age_reconstruction_valid,
        "all_age_count_reconstruction_tolerance_pct":ALL_AGE_RECONSTRUCTION_TOLERANCE_PCT,
        "maximum_absolute_decomposition_closure_error":float(decomp.closure_error.abs().max()),
        "primary_outputs_exclude_percent_metric":True,
        "cause_name_is_schizophrenia":bool(set(burden.cause_name.dropna().astype(str).str.strip())=={"Schizophrenia"}),
        "population_is_official_gbd_2023":bool(pop.population_source.iloc[0]=="official_GBD_2023"),
        "fine_age_burden_validated":fine_age_valid,
        "source_metadata_complete":source_metadata_complete,
        "decomposition_age_group_count":len(decomposition_ages),
        "decomposition_age_bin_sensitivity_flags":int(decomp_age_sensitivity.material_age_bin_sensitivity.sum()),
        "decomposition_path_maximum_shift_pct_of_total_change":float(decomp_path_sensitivity.maximum_path_shift_pct_of_total_change.max()),
        "incidence_supported_age_decomposition_rows":int(len(supported_age_incidence_decomp)),
        "incidence_supported_age_maximum_absolute_closure_error":float(supported_age_incidence_decomp.closure_error.abs().max()) if not supported_age_incidence_decomp.empty else None,
        "segmented_ar1_sensitivity_rows":int(len(ar1_sensitivity)),
        "segmented_ar1_all_iterations_converged":bool(ar1_sensitivity.iterations_converged.all()),
        "segmented_ar1_practical_label_agreement_count":int(ar1_sensitivity.practical_label_agreement.sum()),
        "segmented_ar1_maximum_absolute_aapc_difference_pct_points":float(ar1_sensitivity.ar1_minus_primary_aapc.abs().max()),
        "primary_practical_stability_threshold_pct_per_year":PRACTICAL_STABILITY_THRESHOLD_PCT_PER_YEAR,
        "apc_weighting_maximum_absolute_global_slope_difference_pct_points":float(apc_weighting_sensitivity.equal_minus_population_weighted_global_period_slope.abs().max()),
        "official_nci_results_imported":nci_valid,
        "internal_validation_passed":internal_validation_passed,
        "data_readiness_passed":data_readiness_passed,
        "formal_trend_inference_performed":False,
        "segmented_sensitivity_direction_changes":int((~segmented_sensitivity.direction_stable_vs_primary).sum()),
        "analysis_ready":metadata["analysis_ready"],
    }
    (out/"qa"/"validation_summary.json").write_text(json.dumps(validation,indent=2),encoding="utf-8")
    return {"output":out,"tables":tables,"metadata":metadata}


def parse_args():
    p=argparse.ArgumentParser(description="Build the China-US schizophrenia GBD 2023 analysis results.")
    p.add_argument("--burden-csv",default=DEFAULT_BURDEN,type=Path); p.add_argument("--population-csv",type=Path)
    p.add_argument("--burden-metadata-json",type=Path,help="Production provenance sidecar for the preserved raw burden export.")
    p.add_argument("--population-metadata-json",type=Path,help="Production provenance sidecar for the preserved raw population export.")
    p.add_argument("--population-release",default="GBD 2023"); p.add_argument("--output-dir",default=DEFAULT_OUTPUT,type=Path)
    p.add_argument("--allow-proxy-population",action="store_true",help="Exploratory mode using reconstructed proxy population values.")
    p.add_argument("--nci-results-csv",type=Path,help="Optional normalized NCI output for descriptive curve validation.")
    return p.parse_args()


if __name__=="__main__":
    result=run(parse_args()); print(json.dumps(result["metadata"],indent=2)); print(f"Outputs: {result['output']}")
